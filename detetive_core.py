"""Núcleo do Detetive (sem UI) — usado pela app Django `detetive_fiscal` (modulo Sentinela)."""

import io
import re
import uuid
import zipfile
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor

import pandas as pd


def _xl_col_to_name(col_num: int, col_abs: bool = False) -> str:
    """Índice de coluna 0-based → letras estilo Excel (ex.: 0→A). Mesma regra que xlsxwriter.utility."""
    if col_num < 0:
        return ""
    c = col_num + 1
    col_str = ""
    prefix = "$" if col_abs else ""
    while c:
        remainder = c % 26
        if remainder == 0:
            remainder = 26
        col_str = chr(ord("A") + remainder - 1) + col_str
        c = int((c - 1) / 26)
    return prefix + col_str


def _xl_rowcol_to_cell(
    row_num: int, col_num: int, row_abs: bool = False, col_abs: bool = False
) -> str:
    """Célula A1 a partir de linha/coluna 0-based. Mesma regra que xlsxwriter.utility."""
    if row_num < 0 or col_num < 0:
        return ""
    row_1 = row_num + 1
    row_abs_str = "$" if row_abs else ""
    return _xl_col_to_name(col_num, col_abs) + row_abs_str + str(row_1)


ROTULO_CLIENTE = "SPED Cliente"
ROTULO_NASCEL = "SPED Nascel"
SUF_CLIENTE = "_Cliente"
SUF_NASCEL = "_Nascel"

CHAVES_CANDIDATAS = [
    "CHV_NFE",
    "CHV_CTE",
    "CHV_NFSE",
    "CHAVE_ACESSO",
    "CHV_DOCe",
    "CNPJ",
    "COD_PART",
]

ALIASES_POR_REG = {
    "C100": [("C09", "CHV_NFE")],
    "0150": [("C05", "CNPJ"), ("C02", "COD_PART")],
}

# Bloco E — apuração ICMS (DIFAL/FCP: E200–E250; ST: E300–E316) + apoio (E001–E116).
# FCP (Fundo de Combate à Pobreza): totais em E310/E311; detalhe por NF em C101/D101 (VL_FCP_UF_DEST, …).
_REGS_BLOCO_APUR_ICMS_ST: frozenset[str] = frozenset(
    (
        "C101",
        "D101",
        "E001",
        "E100",
        "E110",
        "E111",
        "E112",
        "E113",
        "E115",
        "E116",
        "E200",
        "E210",
        "E220",
        "E230",
        "E240",
        "E250",
        "E300",
        "E310",
        "E311",
        "E312",
        "E313",
        "E316",
    )
)


def _norm_val(v):
    if pd.isna(v) or v is None:
        return ""
    return str(v).strip()


def _ref_registro_sped_row(row: pd.Series) -> str:
    for base in ("CHV_NFE", "NUM_DOC", "CNPJ", "COD_PART", "CHAVE_ACESSO", "CHV_CTE"):
        for suf in ("", SUF_CLIENTE, SUF_NASCEL):
            col = f"{base}{suf}" if suf else base
            if col in row.index:
                v = _norm_val(row[col])
                if v and v.lower() not in ("nan", "none"):
                    return f"{base} {v}"
    return "—"


def _detalhe_divergencia_campos(row: pd.Series) -> str:
    raw = row.get("COLUNAS_DIVERGENTES", "")
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    partes = []
    for col in str(raw).split(","):
        col = col.strip()
        if not col:
            continue
        if "(" in col:
            partes.append(col)
            continue
        c1, c2 = f"{col}{SUF_CLIENTE}", f"{col}{SUF_NASCEL}"
        if c1 in row.index and c2 in row.index:
            v1, v2 = _norm_val(row[c1]), _norm_val(row[c2])
            if v1 != v2:
                partes.append(f"{col}: Cliente «{v1}» | Nascel «{v2}»")
            else:
                partes.append(col)
        else:
            partes.append(col)
    return "; ".join(partes) if partes else str(raw)


def montar_onde_agir_sped(
    out_div: pd.DataFrame,
    out_so_c: pd.DataFrame,
    out_so_n: pd.DataFrame,
    max_div_detalhe: int = 500,
) -> pd.DataFrame:
    linhas: list[dict] = []
    n = 0

    if out_so_c is not None and not out_so_c.empty and "ABA_SPED" in out_so_c.columns:
        for aba, grp in out_so_c.groupby("ABA_SPED"):
            n += 1
            linhas.append(
                {
                    "#": n,
                    "Prioridade": "Média",
                    "Tipo": f"Só no {ROTULO_CLIENTE}",
                    "Referencia": str(aba),
                    "Acao": f"{len(grp)} linha(s) só no Cliente neste registro. Abrir aba **So_Cliente** (filtrar ABA_SPED = {aba}).",
                }
            )

    if out_so_n is not None and not out_so_n.empty and "ABA_SPED" in out_so_n.columns:
        for aba, grp in out_so_n.groupby("ABA_SPED"):
            n += 1
            linhas.append(
                {
                    "#": n,
                    "Prioridade": "Média",
                    "Tipo": f"Só no {ROTULO_NASCEL}",
                    "Referencia": str(aba),
                    "Acao": f"{len(grp)} linha(s) só na Nascel neste registro. Abrir aba **So_Nascel** (filtro ABA_SPED = {aba}).",
                }
            )

    if out_div is not None and not out_div.empty:
        for i, (_, row) in enumerate(out_div.iterrows()):
            if i >= max_div_detalhe:
                break
            n += 1
            aba = row.get("ABA_SPED", "")
            ref = _ref_registro_sped_row(row)
            det = _detalhe_divergencia_campos(row)
            linhas.append(
                {
                    "#": n,
                    "Prioridade": "Alta",
                    "Tipo": "Mesma chave, campos diferentes",
                    "Referencia": f"{aba} · {ref}",
                    "Acao": det or _norm_val(row.get("COLUNAS_DIVERGENTES", "")),
                }
            )
        overflow = len(out_div) - min(len(out_div), max_div_detalhe)
        if overflow > 0:
            n += 1
            linhas.append(
                {
                    "#": n,
                    "Prioridade": "Alta",
                    "Tipo": "Limite da lista",
                    "Referencia": "—",
                    "Acao": f"Mais {overflow} divergência(s) na aba **Campos_divergentes** (mesmo detalhe por linha).",
                }
            )

    if not linhas:
        linhas.append(
            {
                "#": 1,
                "Prioridade": "—",
                "Tipo": "Sem pendências",
                "Referencia": "—",
                "Acao": "Nada listado: nenhuma diferença nos blocos comparados (ou filtros vazios).",
            }
        )

    return pd.DataFrame(linhas)


# --- Somatório por CFOP (C100 amarra C190) — comparar dois SPEDs ---

_COLS_C190_SOMA = ["VL_OPR", "VL_BC_ICMS", "VL_ICMS", "VL_BC_ICMS_ST", "VL_ICMS_ST", "VL_IPI"]

# Cores nas abas Cliente/Nascel (e D190_*): um único realce para qualquer divergência (valor ou CFOP).
_DESTAQUE_XLSX_FMT_POR_CATEG: dict[str, dict] = {
    "divergencia": {
        "bg_color": "#FCE7F3",
        "font_color": "#831843",
        "border": 1,
        "border_color": "#DB2777",
    },
}
# Identificação da linha (chave, doc., item): neutro — à parte das cores por tipo de valor.
_DESTAQUE_XLSX_FMT_CONTEXTO = {
    "bg_color": "#E7E5E4",
    "font_color": "#292524",
    "border": 1,
    "border_color": "#A8A29E",
}
# Colunas internas a realçar (cinza) quando a linha tem divergência — resto da folha sem cor.
_COL_CTX_DESTAQUE_C190 = ("CHV_NFE", "NUM_DOC", "COD_ITEM", "DESCR_ITEM", "CFOP")
_COL_CTX_DESTAQUE_D190 = ("CHV_DOC", "NUM_DOC", "COD_ITEM", "DESCR_ITEM", "CFOP")
_COL_CFOP_SOM = "CFOP_SPED_soma"

# --- Rótulos em português no Excel (colunas identificáveis; chave de acesso em destaque) ---
_LABEL_ORIGEM_SPED = "Origem do ficheiro (SPED)"
_LABEL_CHAVE_NFE = "Chave de acesso da NF-e (44 dígitos)"
_LABEL_CHAVE_CTE = "Chave de acesso CT-e / documento (44 dígitos)"
_LABEL_CFOP_SOM_TXT = "CFOP em texto (usado no somatório / tabela dinâmica)"

_SOMA_COL_PT = {
    "VL_OPR": "Soma — valor da operação (VL_OPR)",
    "VL_BC_ICMS": "Soma — base de cálculo ICMS (VL_BC_ICMS)",
    "VL_ICMS": "Soma — valor do ICMS (VL_ICMS)",
    "VL_BC_ICMS_ST": "Soma — base ICMS ST (VL_BC_ICMS_ST)",
    "VL_ICMS_ST": "Soma — valor ICMS ST (VL_ICMS_ST)",
    "VL_IPI": "Soma — valor do IPI (VL_IPI)",
}

_CAB_CFOP_EXCEL = ["CFOP"] + [_SOMA_COL_PT[c] for c in _COLS_C190_SOMA]

_COL_VALOR_PT = {
    "VL_OPR": "Valor da operação (VL_OPR)",
    "VL_BC_ICMS": "Base de cálculo ICMS (VL_BC_ICMS)",
    "VL_ICMS": "Valor do ICMS (VL_ICMS)",
    "VL_BC_ICMS_ST": "Base ICMS ST (VL_BC_ICMS_ST)",
    "VL_ICMS_ST": "Valor ICMS ST (VL_ICMS_ST)",
    "VL_IPI": "Valor do IPI (VL_IPI)",
    "VL_RED_BC": "Valor redução da base de cálculo (VL_RED_BC)",
    "VL_DOC": "Valor do documento (VL_DOC)",
}

_COL_DETALHE_C190 = {
    "LADO": _LABEL_ORIGEM_SPED,
    "CHV_NFE": _LABEL_CHAVE_NFE,
    "CHAVE_ACESSO": _LABEL_CHAVE_NFE,
    "CNPJ_EMIT_NFE": "CNPJ do emitente da NF-e (na chave 44)",
    "CFOP": "CFOP",
    _COL_CFOP_SOM: _LABEL_CFOP_SOM_TXT,
    "COD_ITEM": "Código do item (C170)",
    "DESCR_ITEM": "Descrição do item",
    "NUM_DOC": "Número do documento",
    "SER": "Série",
    "COD_PART": "Código do participante",
    "CNPJ": "CNPJ",
    "CST_ICMS": "CST ICMS",
    "REG": "Registro SPED",
}
for _k, _v in _COL_VALOR_PT.items():
    _COL_DETALHE_C190.setdefault(_k, _v)

_COL_DETALHE_D190 = {
    "LADO": _LABEL_ORIGEM_SPED,
    "CHV_DOC": _LABEL_CHAVE_CTE,
    "CHV_CTE": _LABEL_CHAVE_CTE,
    "CHV_NFE": _LABEL_CHAVE_CTE,
    "CFOP": "CFOP",
    _COL_CFOP_SOM: _LABEL_CFOP_SOM_TXT,
    "COD_ITEM": "Código do item",
    "DESCR_ITEM": "Descrição do item",
    "NUM_DOC": "Número do documento",
    "SER": "Série",
    "COD_PART": "Código do participante",
    "CNPJ": "CNPJ",
    "CST_ICMS": "CST ICMS",
    "REG": "Registro SPED",
}
for _k, _v in _COL_VALOR_PT.items():
    _COL_DETALHE_D190.setdefault(_k, _v)

_COL_C100_PLAN = {
    "LADO": _LABEL_ORIGEM_SPED,
    "CHV_NFE": _LABEL_CHAVE_NFE,
    "CHAVE_ACESSO": _LABEL_CHAVE_NFE,
    "NUM_DOC": "Número do documento",
    "SER": "Série",
    "COD_PART": "Código do participante",
    "CNPJ": "CNPJ",
    "VL_DOC": _COL_VALOR_PT["VL_DOC"],
    "IND_OPER": "Indicador operação",
    "IND_EMIT": "Indicador emitente",
    "COD_MOD": "Modelo documento",
    "COD_SIT": "Situação do documento (COD_SIT)",
    "REG": "Registro SPED",
}


def _col_excel_idx(df: pd.DataFrame, interno: str, mapa: dict[str, str]) -> int | None:
    cols = list(df.columns)
    lab = mapa.get(interno, interno)
    if lab in cols:
        return cols.index(lab)
    if interno in cols:
        return cols.index(interno)
    return None


def _to_float_celula_br(v) -> float:
    """Converte valor de célula SPED/Excel: número, texto US, ou BR (ex.: 1.234,56)."""
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, str):
        s = v.strip()
    else:
        try:
            if pd.isna(v):
                return 0.0
        except (TypeError, ValueError):
            pass
        try:
            fv = float(v)
            return 0.0 if fv != fv else fv
        except (TypeError, ValueError):
            s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "-", "–", "—"):
        return 0.0
    num = pd.to_numeric(s, errors="coerce")
    if not pd.isna(num):
        return float(num)
    if "," in s:
        s2 = s.replace(".", "").replace(",", ".")
        try:
            return float(s2)
        except ValueError:
            return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _coerce_colunas_valor_sped(df: pd.DataFrame, colunas: list[str]) -> None:
    for c in colunas:
        if c in df.columns:
            df[c] = df[c].map(_to_float_celula_br)


def _xlsx_largura_coluna_chave(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
    mapa: dict[str, str],
    largura: float = 50,
) -> None:
    """Garante coluna legível para chave de 44 dígitos (NF-e ou CT-e)."""
    if sheet_name not in writer.sheets or df is None or df.empty:
        return
    ws = writer.sheets[sheet_name]
    for interno in ("CHV_NFE", "CHV_DOC"):
        j = _col_excel_idx(df, interno, mapa)
        if j is not None:
            ws.set_column(j, j, largura)
            return


def _xlsx_largura_coluna_cnpj_emitente_nfe(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
) -> None:
    """Largura confortável para a coluna «CNPJ do emitente da NF-e» na aba C190."""
    if sheet_name not in writer.sheets or df is None or df.empty:
        return
    j = _col_excel_idx(df, "CNPJ_EMIT_NFE", _COL_DETALHE_C190)
    if j is not None:
        writer.sheets[sheet_name].set_column(j, j, 18)


def _df_export_c190_planilha(df: pd.DataFrame, lado: str) -> pd.DataFrame:
    """Detalhe C190 com cabeçalhos em português; chave NF-e logo após a origem."""
    x = _df_export_sped_com_cfop_soma(df, lado)
    if x.empty:
        return x
    out = x.copy()
    if "CHV_NFE" not in out.columns:
        out["CHV_NFE"] = ""
    if "CHAVE_ACESSO" in x.columns:
        out["CHV_NFE"] = out["CHV_NFE"].replace("", pd.NA).fillna(x["CHAVE_ACESSO"])
    out["CHV_NFE"] = out["CHV_NFE"].map(
        lambda v: _digits_chave(v) if pd.notna(v) and str(v).strip() else ""
    )
    if "COD_ITEM" not in out.columns:
        out["COD_ITEM"] = ""
    if "DESCR_ITEM" not in out.columns:
        out["DESCR_ITEM"] = ""
    out["CNPJ_EMIT_NFE"] = out["CHV_NFE"].map(_cnpj_emitente_de_chave_nfe_44)
    if "IND_EMIT" in out.columns and "CNPJ" in out.columns:
        ie = out["IND_EMIT"].astype(str).str.strip()
        cn = out["CNPJ"].map(lambda v: re.sub(r"\D", "", _norm_val(v))[:14])
        fb = out["CNPJ_EMIT_NFE"].eq("") & ie.eq("1") & cn.str.len().eq(14)
        out.loc[fb, "CNPJ_EMIT_NFE"] = cn.loc[fb]
    pref = [
        "LADO",
        "CHV_NFE",
        "CNPJ_EMIT_NFE",
        "NUM_DOC",
        "COD_ITEM",
        "DESCR_ITEM",
        "CFOP",
        _COL_CFOP_SOM,
        "CST_ICMS",
    ]
    vl_rest = [c for c in _COLS_C190_SOMA if c in out.columns and c not in pref]
    _coerce_colunas_valor_sped(out, vl_rest)
    if "VL_RED_BC" in out.columns:
        _coerce_colunas_valor_sped(out, ["VL_RED_BC"])
    outros = [c for c in out.columns if c not in pref + vl_rest]
    ordered = [c for c in pref if c in out.columns] + vl_rest + outros
    out = out[ordered]
    ren = {k: v for k, v in _COL_DETALHE_C190.items() if k in out.columns}
    return out.rename(columns=ren)


def _df_export_d190_planilha(df: pd.DataFrame, lado: str) -> pd.DataFrame:
    x = _df_export_sped_com_cfop_soma(df, lado)
    if x.empty:
        return x
    out = x.copy()
    if "CHV_DOC" not in out.columns:
        out["CHV_DOC"] = ""
    if "CHV_CTE" in x.columns:
        out["CHV_DOC"] = out["CHV_DOC"].replace("", pd.NA).fillna(x["CHV_CTE"])
    if "CHV_NFE" in x.columns and out["CHV_DOC"].eq("").all():
        out["CHV_DOC"] = out["CHV_DOC"].mask(out["CHV_DOC"].eq(""), x["CHV_NFE"])
    out["CHV_DOC"] = out["CHV_DOC"].map(
        lambda v: _digits_chave(v) if pd.notna(v) and str(v).strip() else ""
    )
    if "COD_ITEM" not in out.columns:
        out["COD_ITEM"] = ""
    if "DESCR_ITEM" not in out.columns:
        out["DESCR_ITEM"] = ""
    pref = ["LADO", "CHV_DOC", "NUM_DOC", "COD_ITEM", "DESCR_ITEM", "CFOP", _COL_CFOP_SOM, "CST_ICMS"]
    vl_rest = [c for c in _COLS_C190_SOMA if c in out.columns and c not in pref]
    if "VL_RED_BC" in out.columns and "VL_RED_BC" not in pref + vl_rest:
        vl_rest.append("VL_RED_BC")
    _coerce_colunas_valor_sped(out, vl_rest)
    outros = [c for c in out.columns if c not in pref + vl_rest]
    ordered = [c for c in pref if c in out.columns] + vl_rest + outros
    out = out[ordered]
    ren = {k: v for k, v in _COL_DETALHE_D190.items() if k in out.columns}
    return out.rename(columns=ren)


def _df_c100_planilha(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df if df is not None else pd.DataFrame()
    out = df.copy()
    if "CHV_NFE" not in out.columns:
        out["CHV_NFE"] = ""
    if "CHAVE_ACESSO" in df.columns:
        out["CHV_NFE"] = out["CHV_NFE"].replace("", pd.NA).fillna(df["CHAVE_ACESSO"])
    out["CHV_NFE"] = out["CHV_NFE"].map(
        lambda v: _digits_chave(v) if pd.notna(v) and str(v).strip() else ""
    )
    if "VL_DOC" in out.columns:
        _coerce_colunas_valor_sped(out, ["VL_DOC"])
    pref = ["LADO", "CHV_NFE", "NUM_DOC", "SER", "CNPJ", "COD_PART", "VL_DOC", "COD_MOD"]
    if "COD_SIT" in out.columns and "COD_SIT" not in pref:
        pref.append("COD_SIT")
    outros = [c for c in out.columns if c not in pref]
    ordered = [c for c in pref if c in out.columns] + outros
    out = out[ordered]
    ren = {k: v for k, v in _COL_C100_PLAN.items() if k in out.columns}
    return out.rename(columns=ren)


def _comparativo_tabela3_para_excel(comp: pd.DataFrame) -> pd.DataFrame:
    if comp.empty:
        return comp
    m = {
        "CFOP SPED": "CFOP",
        "Presenca_CFOP": "Presença do CFOP (Cliente / Nascel / só um lado)",
        "Divergente": "Divergente (SIM ou NÃO)",
        "Observação": "Observação (preencher à mão)",
    }
    for col in _COLS_C190_SOMA:
        ant = f"Dif. {col} (Cli−Nascel)"
        leg = _COL_VALOR_PT.get(col, col)
        m[ant] = f"Diferença — {leg} (Cliente − Nascel)"
    return comp.rename(columns=m)


# ONDE_AGIR: uma linha por nota; Status = COD_SIT do C100/D100 (quando existir).
_ORDEM_COLS_ONDE_AGIR_CFOP = (
    "#",
    "Nota",
    "Chave_acesso",
    "CFOP_CLIENTE",
    "CFOP_NASCEL",
    "Status",
    "O_que",
)

# C100/D100 — COD_SIT (EFD ICMS/IPI, situação do documento fiscal)
_COD_SIT_DESCR: dict[str, str] = {
    "00": "Autorizada / regular",
    "01": "Regular extemporânea",
    "02": "Cancelada",
    "03": "Cancelada extemporânea",
    "04": "Denegada",
    "05": "Numeração inutilizada",
    "06": "Documento complementar",
    "07": "Complementar extemporâneo",
    "08": "Regime especial / norma específica",
}

# Etiquetas na coluna «O que está diferente» (uma nota pode juntar várias).
_TAG_ONDE_CFOP_DIVERGE = "CFOP diverge"
_TAG_ONDE_SO_CLIENTE = "Só Cliente"
_TAG_ONDE_SO_NASCEL = "Só Nascel"
_TAG_ONDE_METRICA: dict[str, str] = {
    "VL_OPR": "Operação diverge",
    "VL_BC_ICMS": "Base ICMS diverge",
    "VL_ICMS": "ICMS diverge",
    "VL_BC_ICMS_ST": "Base ST diverge",
    "VL_ICMS_ST": "ICMS ST diverge",
    "VL_IPI": "IPI diverge",
}
_ORDEM_TAGS_ONDE: tuple[str, ...] = (
    _TAG_ONDE_CFOP_DIVERGE,
    _TAG_ONDE_SO_CLIENTE,
    _TAG_ONDE_SO_NASCEL,
) + tuple(_TAG_ONDE_METRICA[c] for c in _COLS_C190_SOMA)
_IDX_TAG_ONDE: dict[str, int] = {t: i for i, t in enumerate(_ORDEM_TAGS_ONDE)}


def _norm_cod_sit_sped(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(_norm_val(v)).strip()
    if not s or s.lower() in ("nan", "none"):
        return ""
    if s.isdigit():
        return s.zfill(2)[:2]
    try:
        n = int(float(s.replace(",", ".")))
        return str(n).zfill(2)[:2]
    except (TypeError, ValueError):
        return s[:2] if len(s) >= 2 else s


def _descr_cod_sit(cod: str) -> str:
    c = _norm_cod_sit_sped(cod)
    if not c:
        return "—"
    return _COD_SIT_DESCR.get(c, f"Código {c}")


def _mapa_status_por_chave_doc(df_head: pd.DataFrame | None) -> dict[str, str]:
    """
    Chave 44 dígitos → texto da situação (COD_SIT do C100 ou D100).
    """
    if df_head is None or df_head.empty or "COD_SIT" not in df_head.columns:
        return {}
    ch_col = None
    for c in ("CHV_NFE", "CHV_DOC", "CHV_CTE"):
        if c in df_head.columns:
            ch_col = c
            break
    if ch_col is None:
        return {}
    out: dict[str, str] = {}
    for _, row in df_head.iterrows():
        ch = _digits_chave(row.get(ch_col, ""))
        if len(ch) < 10:
            continue
        if ch in out:
            continue
        c = _norm_cod_sit_sped(row.get("COD_SIT"))
        if not c:
            continue
        out[ch] = _descr_cod_sit(c)
    return out


def _status_onde_agir_linha(ch_key: str, map_cli: dict[str, str], map_nas: dict[str, str]) -> str:
    sc = map_cli.get(ch_key, "")
    sn = map_nas.get(ch_key, "")
    if sc and sn and sc != sn:
        return f"Cliente: {sc} | Nascel: {sn}"
    if sc:
        return sc
    if sn:
        return sn
    return "—"


def _onde_agir_ref_nota(num_doc: str, chave44: str) -> str:
    nd = (num_doc or "").strip()
    if nd:
        return nd
    ch = _digits_chave(chave44)
    if len(ch) >= 8:
        return f"Chave …{ch[-8:]}"
    return "—"


def _formatar_conjunto_cfop_excel(s: set[str]) -> str:
    """Lista de CFOP por lado, ordenada, para a aba Onde agir."""
    if not s:
        return "—"
    return "; ".join(sorted(s, key=lambda x: (len(str(x)), str(x))))


def _onde_agir_tags_ordenadas(tags: set[str]) -> str:
    if not tags:
        return ""
    return "; ".join(sorted(tags, key=lambda t: _IDX_TAG_ONDE.get(t, 99)))


def _onde_agir_cfop_para_excel(onde: pd.DataFrame) -> pd.DataFrame:
    if onde.empty:
        return onde
    m = {
        "#": "N.º",
        "Nota": "Nota",
        "Chave_acesso": "Chave de acesso",
        "CFOP_CLIENTE": "CFOP Cliente",
        "CFOP_NASCEL": "CFOP Nascel",
        "Status": "Status no SPED",
        "O_que": "Resumo (tipo de divergência, por nota)",
        "Tipo": "Tipo",
        "Referencia": "Sobre",
    }
    ordem = [c for c in _ORDEM_COLS_ONDE_AGIR_CFOP if c in onde.columns]
    ordem += [c for c in onde.columns if c not in ordem]
    onde = onde.loc[:, ordem].copy()
    return onde.rename(columns={c: m.get(c, c) for c in onde.columns})


# Identidade visual Detetive no Excel — **pink + laranja** (sem verde/azul do tema Office).
_FMT_XLSX_TITULO_BLOCO_PINK = {
    "bold": True,
    "bg_color": "#FBCFE8",
    "font_color": "#831843",
    "font_size": 11,
}
_FMT_XLSX_TITULO_BLOCO_LARANJA = {
    "bold": True,
    "bg_color": "#FED7AA",
    "font_color": "#9A3412",
    "font_size": 11,
}
# Compat: somatório único e títulos que não alternam cor.
_FMT_XLSX_TITULO_BLOCO = _FMT_XLSX_TITULO_BLOCO_PINK

# Light 1 = branco / cinza muito claro — não segue o «accent» do tema (evita verde/azul aleatório).
_XLSX_TABLE_STYLE_DINAMICA = "Table Style Light 1"

# ListObject + autofiltro do Excel deixa células «estruturadas» e incomoda quem quer editar à mão.
# False = não cria tabela — folha livre (podes filtrar/ordenar com o autofiltro clássico se quiseres).
_XLSX_USAR_TABELA_ESTRUTURADA = False

_ABA_PINK_HEX = "#DB2777"
_ABA_LARANJA_HEX = "#EA580C"


def _xlsx_escrever_legenda_destaque_cfop(writer: pd.ExcelWriter) -> None:
    """
    Folha única: realces nas abas de detalhe (Cliente, Nascel, D190_*) + como usar a aba «Onde agir».
    Só cria uma vez por workbook (idempotente).
    """
    nome = _ABA_LEGENDA_CORES[:31]
    wb = writer.book
    if wb.get_worksheet_by_name(nome):
        return
    ws = wb.add_worksheet(nome)
    fmt_tit = wb.add_format(
        {"bold": True, "font_size": 12, "bg_color": "#FCE7F3", "font_color": "#831843"}
    )
    fmt_hdr = wb.add_format({"bold": True, "bottom": 1})
    fmt_txt = wb.add_format({"text_wrap": True, "valign": "vcenter"})

    def _fmt_num(cat: str) -> object:
        d = dict(_DESTAQUE_XLSX_FMT_POR_CATEG[cat])
        d["num_format"] = "#,##0.00"
        return wb.add_format(d)

    fmt_ctx = wb.add_format({**_DESTAQUE_XLSX_FMT_CONTEXTO, "num_format": "@"})
    fmt_div = _fmt_num("divergencia")
    fmt_div_txt = wb.add_format(
        {**dict(_DESTAQUE_XLSX_FMT_POR_CATEG["divergencia"]), "num_format": "@"}
    )

    linhas: list[tuple[object, str | float, bool, str]] = [
        (
            fmt_ctx,
            "Chave / doc.",
            True,
            "Cinza: identificação da linha (chave, documento, item, CFOP). Não significa erro por si só.",
        ),
        (
            fmt_div,
            1234.56,
            False,
            "Rosa: célula com diferença entre os dois SPED (qualquer valor VL_* ou CFOP). A mesma cor para todos os casos.",
        ),
        (
            fmt_div_txt,
            "5102",
            True,
            "Rosa (texto): mesma regra na coluna CFOP quando houver divergência.",
        ),
    ]

    ws.merge_range(0, 0, 0, 1, "Legenda — realces e aba «Onde agir»", fmt_tit)
    ws.write(1, 0, "Exemplo", fmt_hdr)
    ws.write(1, 1, "Significado", fmt_hdr)
    r = 2
    for fmt_am, val, as_str, desc in linhas:
        if as_str:
            ws.write_string(r, 0, str(val), fmt_am)
        else:
            ws.write_number(r, 0, float(val), fmt_am)
        ws.write_string(r, 1, desc, fmt_txt)
        r += 1
    nota = (
        "Onde agir: use a aba ONDE_AGIR (ou ONDE_AGIR_NF_e / ONDE_AGIR_CTe) — uma linha por nota, "
        "coluna «Resumo» com o tipo de pendência (CFOP, ICMS, só um SPED, etc.). "
        "Abra o detalhe nas folhas Cliente/Nascel filtrando pela chave ou pelo n.º do documento."
    )
    ws.merge_range(r, 0, r, 1, nota, fmt_txt)
    ws.set_row(r, 48)
    ws.set_column(0, 0, 22)
    ws.set_column(1, 1, 88)


def _xlsx_escrever_blocos_apur_difal_st(
    writer: pd.ExcelWriter,
    bl_c: dict[str, pd.DataFrame],
    bl_n: dict[str, pd.DataFrame] | None = None,
) -> None:
    """
    Folhas opcionais: Bloco E (DIFAL, **FCP**, ST), mais **C101/D101** (complemento por documento com FCP/DIFAL por UF).
    Dois SPED: sufixo _Cliente / _Nascel; um SPED: nome do registro (ex.: E310, C101).
    """
    usados: set[str] = set(writer.sheets)

    def _nome_aba(base: str) -> str:
        n = base[:31]
        k = 0
        while n in usados:
            k += 1
            suf = f"_{k}"
            n = (base[: 31 - len(suf)] + suf)[:31]
        usados.add(n)
        return n

    def _gravar(df: pd.DataFrame, nome_base: str) -> None:
        if df is None or df.empty:
            return
        nm = _nome_aba(nome_base)
        df.to_excel(writer, index=False, sheet_name=nm)
        ws = writer.sheets[nm]
        ncol = max(1, len(df.columns)) - 1
        ws.set_column(0, min(ncol, 60), 14)

    if bl_n is None:
        for reg in sorted(_REGS_BLOCO_APUR_ICMS_ST):
            df = bl_c.get(reg)
            if isinstance(df, pd.DataFrame):
                _gravar(df, reg)
        return

    for reg in sorted(_REGS_BLOCO_APUR_ICMS_ST):
        dfc = bl_c.get(reg)
        if isinstance(dfc, pd.DataFrame):
            _gravar(dfc, f"{reg}_Cliente")
        dfn = bl_n.get(reg)
        if isinstance(dfn, pd.DataFrame):
            _gravar(dfn, f"{reg}_Nascel")


def _xlsx_aplicar_abas_rosa(writer: pd.ExcelWriter) -> None:
    """Separadores das folhas alternando pink e laranja."""
    for i, (_nm, ws) in enumerate(writer.sheets.items()):
        try:
            ws.set_tab_color(_ABA_PINK_HEX if i % 2 == 0 else _ABA_LARANJA_HEX)
        except Exception:
            pass


_ABA_LEGENDA_CORES = "Legenda_cores"

# Ordem das abas no Excel comparativo: comparativo → legenda de cores → onde agir → detalhe.
_ORDEM_ABAS_COMP_C190 = [
    "Comparativo_CFOP_C190",
    _ABA_LEGENDA_CORES,
    "ONDE_AGIR",
    "Cliente",
    "Nascel",
]
_ORDEM_ABAS_COMP_D190 = [
    "Comparativo_CFOP_D190",
    _ABA_LEGENDA_CORES,
    "ONDE_AGIR",
    "D100_Cliente",
    "D100_Nascel",
    "D190_Cliente",
    "D190_Nascel",
]
_ORDEM_ABAS_COMP_NFE_CTE = [
    "Comparativo_CFOP_C190",
    "Comparativo_CFOP_D190",
    _ABA_LEGENDA_CORES,
    "ONDE_AGIR_NF_e",
    "ONDE_AGIR_CTe",
    "Cliente",
    "Nascel",
    "D100_Cliente",
    "D100_Nascel",
    "D190_Cliente",
    "D190_Nascel",
]


def _xlsx_liberar_folhas_e_autofiltro(wb) -> None:
    """Folha sem proteção; autofiltro no bloco usado — Excel abre pronto para editar e filtrar."""
    from openpyxl.chartsheet.chartsheet import Chartsheet
    from openpyxl.worksheet.protection import SheetProtection

    for ws in wb.worksheets:
        if isinstance(ws, Chartsheet):
            continue
        ws.protection = SheetProtection(sheet=False)
        try:
            dim = ws.calculate_dimension()
            if not dim or ":" not in dim:
                continue
            c1, c2 = dim.split(":", 1)
            if c1 == c2 and ws[c1].value is None:
                continue
            ws.auto_filter.ref = dim
        except Exception:
            pass


def _xlsx_workbook_posprocesso_liberar(buf: io.BytesIO) -> io.BytesIO:
    """Regrava o .xlsx sem proteção de folhas e com autofiltro no intervalo de dados."""
    from openpyxl import load_workbook

    buf.seek(0)
    wb = load_workbook(buf)
    _xlsx_liberar_folhas_e_autofiltro(wb)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _xlsx_reordenar_abas_por_lista(buf: io.BytesIO, ordem_desejada: list[str]) -> io.BytesIO:
    """Reordena folhas no .xlsx já gravado (openpyxl), mantendo o resto no fim."""
    from openpyxl import load_workbook
    from openpyxl.styles.colors import Color

    buf.seek(0)
    wb = load_workbook(buf)
    presentes = list(wb.sheetnames)
    nova = [n for n in ordem_desejada if n in presentes]
    for n in presentes:
        if n not in nova:
            nova.append(n)
    por_titulo = {ws.title: ws for ws in wb.worksheets}
    wb._sheets = [por_titulo[t] for t in nova]
    # Tabs pink/laranja na ordem **final** (após o xlsxwriter ter gravado noutra ordem).
    pink, lar = "FFDB2777", "FFEA580C"
    for i, ws in enumerate(wb.worksheets):
        try:
            ws.sheet_properties.tabColor = Color(rgb=pink if i % 2 == 0 else lar)
        except Exception:
            pass
    _xlsx_liberar_folhas_e_autofiltro(wb)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _xlsx_add_tabela_estilo_dinamica(
    ws,
    first_row: int,
    last_row: int,
    first_col: int,
    last_col: int,
    headers: list[str] | tuple[str, ...] | None = None,
) -> None:
    """
    Envolve o intervalo numa tabela com filtro. Sem «columns» com header, o Excel
    mostra «Column1», «Column2»… por cima dos dados — daí os rótulos explícitos.
    """
    if not _XLSX_USAR_TABELA_ESTRUTURADA:
        return
    if last_row < first_row or last_col < first_col:
        return
    ncols = last_col - first_col + 1
    opts: dict = {
        "style": _XLSX_TABLE_STYLE_DINAMICA,
        "autofilter": True,
        "name": f"T_{uuid.uuid4().hex[:12]}",
    }
    if headers is not None:
        hlist = [
            str(headers[i]) if i < len(headers) else f"Coluna_{i + 1}"
            for i in range(ncols)
        ]
        opts["columns"] = [{"header": h} for h in hlist]
    ws.add_table(first_row, first_col, last_row, last_col, opts)


def _xlsx_tabela_sobre_df_escrito(ws, df: pd.DataFrame) -> None:
    """DataFrame já escrito com header na linha 0: cabeçalho + len(df) linhas de dados."""
    if df is None or len(df.columns) < 1:
        return
    n_linhas_dados = len(df)
    if n_linhas_dados < 0:
        return
    n_cols = len(df.columns)
    last_r = n_linhas_dados  # header row 0 + dados 1..n_linhas_dados
    _xlsx_add_tabela_estilo_dinamica(
        ws, 0, last_r, 0, n_cols - 1, list(df.columns)
    )


def _norm_cfop_sped(v) -> str:
    t = _norm_val(v).replace(".0", "").strip()
    if not t or t.lower() in ("nan", "none"):
        return "(vazio)"
    return t


def _bucket_regs_sped_por_texto(texto: str, regs: frozenset[str]) -> dict[str, pd.DataFrame]:
    """Extrai linhas |REG|… do texto EFD apenas para os registros pedidos (ex.: E210, E300)."""
    buckets: dict[str, list[dict]] = defaultdict(list)
    for line in texto.splitlines():
        parts = _partes_linha_sped(line)
        if not parts:
            continue
        reg = parts[0].strip().upper()
        if reg not in regs:
            continue
        row = {"REG": reg}
        for i, val in enumerate(parts[1:], start=2):
            row[f"C{i:02d}"] = val
        buckets[reg].append(row)
    out: dict[str, pd.DataFrame] = {}
    for reg, rows in buckets.items():
        if rows:
            df = pd.DataFrame(rows).fillna("").astype(str)
            out[reg] = _aplicar_aliases_reg(reg, df)
    return out


def _df_sped_lib(d: dict, reg: str) -> pd.DataFrame:
    x = d.get(reg)
    return x.copy() if isinstance(x, pd.DataFrame) else pd.DataFrame()


def _carregar_abas_regs_excel(file_obj, regs: frozenset[str]) -> dict[str, pd.DataFrame]:
    """Lê do .xlsx/.xlsm abas cujo nome (parte antes de « - ») coincide com um registro EFD."""
    out: dict[str, pd.DataFrame] = {}
    file_obj.seek(0)
    try:
        xl = pd.ExcelFile(file_obj)
    except Exception:
        file_obj.seek(0)
        return out
    for sh in xl.sheet_names:
        key = _reg_id_de_nome_aba(sh)
        if key not in regs:
            continue
        try:
            file_obj.seek(0)
            df = pd.read_excel(file_obj, sheet_name=sh, dtype=str)
        except Exception:
            continue
        df = df.fillna("").astype(str)
        df = _aplicar_aliases_reg(key, df)
        if key in out:
            out[key] = pd.concat([out[key], df], ignore_index=True)
        else:
            out[key] = df
    file_obj.seek(0)
    return out


def _carregar_blocos_txt_via_spedlib(file_obj, filename: str) -> dict[str, pd.DataFrame]:
    """
    Leitura .txt: remove assinatura (|9999|) + EFDReader (spedlib) + reforço posicional
    de VL_DOC / totais (split | como no Guia), para não deslocar valores quando o arquivo
    tem quantidade de campos diferente do layout fixo do reader.
    """
    import os
    import tempfile

    from spedlib.efd_reader import EFDReader
    from spedlib.utils import remove_efd_signature

    file_obj.seek(0)
    raw = file_obj.read()
    if isinstance(raw, str):
        raw = raw.encode("latin-1", errors="replace")

    path_in = None
    path_clean = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".efd.txt") as tmp_in:
            path_in = tmp_in.name
            tmp_in.write(raw)
        path_clean = path_in + ".sem_assinatura.txt"
        remove_efd_signature(path_in, path_clean, encoding="latin-1")
        reader = EFDReader(encoding="latin-1")
        reader.read_file(path_clean)
        d = reader.data
        blocos = {
            "C100": d["C100"].copy(),
            "C190": d["C190"].copy(),
            "C170": d["C170"].copy(),
            "0200": d["0200"].copy(),
            "D100": d["D100"].copy(),
            "D190": d["D190"].copy(),
        }
        texto_clean = ""
        try:
            with open(path_clean, encoding="latin-1") as tf:
                texto_clean = tf.read()
            _reforcar_valores_sped_txt_por_posicao(texto_clean, blocos)
        except OSError:
            pass
        if texto_clean:
            e_txt = _bucket_regs_sped_por_texto(texto_clean, _REGS_BLOCO_APUR_ICMS_ST)
            for reg in sorted(_REGS_BLOCO_APUR_ICMS_ST):
                lib_df = _df_sped_lib(d, reg)
                txt_df = e_txt.get(reg, pd.DataFrame())
                chosen = lib_df if not lib_df.empty else txt_df
                if not chosen.empty:
                    blocos[reg] = chosen
        return blocos
    finally:
        for p in (path_in, path_clean):
            if p and os.path.isfile(p):
                try:
                    os.unlink(p)
                except OSError:
                    pass
        file_obj.seek(0)


def carregar_blocos_sped_completos(file_obj, filename: str) -> dict[str, pd.DataFrame]:
    """C100, C190, D100, D190; mais C101/D101 (FCP por NF) e Bloco E (DIFAL/FCP/ST) quando existirem — .txt ou .xlsx."""
    import confronto_gerencial as cg

    fn = (filename or "").lower()
    file_obj.seek(0)
    if fn.endswith(".txt"):
        blocos = _carregar_blocos_txt_via_spedlib(file_obj, filename)
    else:
        mapa = cg.carregar_mapa_sped(file_obj, filename)
        file_obj.seek(0)
        c100, c190 = cg._df_c100_c190_de_mapa_sped(mapa)
        d100, d190 = cg._df_d100_d190_de_mapa_sped(mapa)
        blocos = {
            "C100": c100.copy(),
            "C190": c190.copy(),
            "C170": pd.DataFrame(),
            "0200": pd.DataFrame(),
            "D100": d100.copy(),
            "D190": d190.copy(),
        }
        if fn.endswith((".xlsx", ".xlsm")):
            for reg, df_e in _carregar_abas_regs_excel(file_obj, _REGS_BLOCO_APUR_ICMS_ST).items():
                if not df_e.empty:
                    blocos[reg] = df_e.copy()
    file_obj.seek(0)
    return blocos


def carregar_df_c190_de_arquivo(file_obj, filename: str) -> pd.DataFrame:
    """C190 já ligado ao C100 (hierarquia .txt ou merge no Excel)."""
    bl = carregar_blocos_sped_completos(file_obj, filename)
    return bl["C190"].copy()


def _agg_cfop_sem_total(df_c190: pd.DataFrame) -> pd.DataFrame:
    if df_c190.empty or "CFOP" not in df_c190.columns:
        return pd.DataFrame(columns=["CFOP SPED"] + _COLS_C190_SOMA)
    g = df_c190.copy()
    for c in _COLS_C190_SOMA:
        if c not in g.columns:
            g[c] = 0.0
        else:
            g[c] = g[c].map(_to_float_celula_br)
    g["CFOP SPED"] = g["CFOP"].map(_norm_cfop_sped)
    agg = g.groupby("CFOP SPED", dropna=False)[_COLS_C190_SOMA].sum().reset_index()
    return agg.round(2)


def tabela_somatorio_cfop_estilo_excel(df_c190: pd.DataFrame) -> pd.DataFrame:
    """Tabela dinâmica por CFOP com cabeçalhos em português (CFOP + Soma — …)."""
    agg = _agg_cfop_sem_total(df_c190)
    if agg.empty:
        return pd.DataFrame(columns=_CAB_CFOP_EXCEL)
    ren: dict[str, str] = {"CFOP SPED": "CFOP"}
    for k in _COLS_C190_SOMA:
        if k in agg.columns:
            ren[k] = _SOMA_COL_PT[k]
    out = agg.rename(columns=ren)
    tot = out.drop(columns=["CFOP"]).sum(numeric_only=True)
    total_row: dict = {"CFOP": "Total Geral"}
    for c in out.columns:
        if c != "CFOP":
            total_row[c] = round(float(tot.get(c, 0.0)), 2)
    out = pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)
    return out


def _comparativo_cfop_merge(agg_c: pd.DataFrame, agg_n: pd.DataFrame) -> pd.DataFrame:
    if agg_c.empty and agg_n.empty:
        return pd.DataFrame()
    set_c = set(agg_c["CFOP SPED"]) if not agg_c.empty else set()
    set_n = set(agg_n["CFOP SPED"]) if not agg_n.empty else set()
    m = agg_c.merge(agg_n, on="CFOP SPED", how="outer", suffixes=("_Cliente", "_Nascel")).fillna(0.0)
    linhas = []
    tol = 0.02
    for _, row in m.iterrows():
        cf = row["CFOP SPED"]
        if cf in set_c and cf in set_n:
            pres = "Ambos"
        elif cf in set_c:
            pres = f"Só {ROTULO_CLIENTE}"
        else:
            pres = f"Só {ROTULO_NASCEL}"
        r = {"CFOP SPED": cf, "Presenca_CFOP": pres}
        div = pres != "Ambos"
        for col in _COLS_C190_SOMA:
            vc = float(row.get(f"{col}_Cliente", 0.0))
            vn = float(row.get(f"{col}_Nascel", 0.0))
            d = round(vc - vn, 2)
            # Só diferença (Cliente − Nascel); totais estão nas Tabelas 1 e 2
            r[f"Dif. {col} (Cli−Nascel)"] = d
            if abs(d) >= tol:
                div = True
        r["Divergente"] = "SIM" if div else "NÃO"
        r["Observação"] = ""
        linhas.append(r)
    out = pd.DataFrame(linhas)
    if not out.empty:
        out = out.sort_values("CFOP SPED")
    return out


def _sheet_ref_excel(nome: str) -> str:
    n = (nome or "")[:31].replace("'", "''")
    return f"'{n}'"


def _refs_sumifs(
    sheet: str,
    col_sum_letter: str,
    col_crit_letter: str,
    excel_row_first: int,
    excel_row_last: int,
) -> tuple[str, str]:
    esc = _sheet_ref_excel(sheet)
    s = f"{esc}!${col_sum_letter}${excel_row_first}:${col_sum_letter}${excel_row_last}"
    c = f"{esc}!${col_crit_letter}${excel_row_first}:${col_crit_letter}${excel_row_last}"
    return s, c


def _pode_formulas_somatorio(
    df: pd.DataFrame,
    mapa: dict[str, str] | None = None,
) -> bool:
    if df is None or df.empty:
        return False
    m = mapa or _COL_DETALHE_C190
    if _col_excel_idx(df, _COL_CFOP_SOM, m) is None:
        return False
    return all(_col_excel_idx(df, c, m) is not None for c in _COLS_C190_SOMA)


def _letras_colunas_soma(
    df: pd.DataFrame,
    mapa: dict[str, str] | None = None,
) -> dict[str, str] | None:
    m = mapa or _COL_DETALHE_C190
    if not _pode_formulas_somatorio(df, m):
        return None
    out: dict[str, str] = {}
    for c in _COLS_C190_SOMA:
        j = _col_excel_idx(df, c, m)
        if j is None:
            return None
        out[c] = _xl_col_to_name(j)
    jc = _col_excel_idx(df, _COL_CFOP_SOM, m)
    if jc is None:
        return None
    out["_crit"] = _xl_col_to_name(jc)
    return out


def _xlsx_reforcar_detalhe_texto_e_numero(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
    mapa: dict[str, str] | None = None,
) -> None:
    """Coluna de CFOP texto @; colunas de valor como número (SUMIFS)."""
    m = mapa or _COL_DETALHE_C190
    if df is None or df.empty:
        return
    ix_cf = _col_excel_idx(df, _COL_CFOP_SOM, m)
    if ix_cf is None:
        return
    ws = writer.sheets[sheet_name]
    wb = writer.book
    fmt_txt = wb.add_format({"num_format": "@"})
    fmt_num = wb.add_format({"num_format": "#,##0.00"})
    n = len(df)
    for r in range(n):
        ws.write_string(r + 1, ix_cf, str(df.iloc[r, ix_cf]), fmt_txt)
    for cname in _COLS_C190_SOMA:
        j = _col_excel_idx(df, cname, m)
        if j is None:
            continue
        for r in range(n):
            raw = df.iloc[r, j]
            num = _to_float_celula_br(raw)
            ws.write_number(r + 1, j, num, fmt_num)


def _xlsx_reforcar_cfop_sped_coluna_a_texto(
    ws,
    tab: pd.DataFrame,
    start_header_row: int,
    fmt_txt,
) -> None:
    """Coluna A em formato texto @; CFOP igual ao detalhe (norm) exceto linha Total Geral."""
    if tab.empty:
        return
    n = len(tab)
    col_cfop = "CFOP" if "CFOP" in tab.columns else "CFOP SPED"
    for i in range(n):
        xrow = start_header_row + 1 + i
        cf = tab.iloc[i][col_cfop]
        if str(cf).strip() == "Total Geral":
            txt = "Total Geral"
        else:
            txt = str(_norm_cfop_sped(cf))
        ws.write_string(xrow, 0, txt, fmt_txt)


def _sobrescrever_somatorio_com_formulas(
    ws,
    tab: pd.DataFrame,
    start_header_row: int,
    sheet_dados: str,
    letras: dict[str, str],
    excel_row_dados_ini: int,
    excel_row_dados_fim: int,
    fmt_num,
) -> None:
    if tab.empty or len(tab) < 2:
        return
    hdr = list(tab.columns)
    n = len(tab)
    for i in range(n - 1):
        xrow = start_header_row + 1 + i
        crit_cell = _xl_rowcol_to_cell(xrow, 0, row_abs=True, col_abs=True)
        row_vals = tab.iloc[i]
        for py_col in _COLS_C190_SOMA:
            col_excel = _SOMA_COL_PT[py_col]
            if col_excel not in hdr:
                continue
            j = hdr.index(col_excel)
            sum_rng, crit_rng = _refs_sumifs(
                sheet_dados,
                letras[py_col],
                letras["_crit"],
                excel_row_dados_ini,
                excel_row_dados_fim,
            )
            fmla = f"=SUMIFS({sum_rng},{crit_rng},{crit_cell})"
            try:
                cached = float(row_vals[col_excel])
            except (TypeError, ValueError):
                cached = 0.0
            ws.write_formula(xrow, j, fmla, fmt_num, cached)
    xrow_tot = start_header_row + 1 + (n - 1)
    r_first = start_header_row + 2
    r_last = start_header_row + n
    tot_vals = tab.iloc[n - 1]
    for py_col in _COLS_C190_SOMA:
        col_excel = _SOMA_COL_PT[py_col]
        if col_excel not in hdr:
            continue
        j = hdr.index(col_excel)
        cl = _xl_col_to_name(j)
        # 9 = SOMA; ignora linhas ocultas pelo filtro automático nesta folha
        ftot = f"=SUBTOTAL(9,{cl}{r_first}:{cl}{r_last})"
        try:
            cached_t = float(tot_vals[col_excel])
        except (TypeError, ValueError):
            cached_t = 0.0
        ws.write_formula(xrow_tot, j, ftot, fmt_num, cached_t)


def _escrever_comparativo_cfop_tres_tabelas(
    writer: pd.ExcelWriter,
    tab_cli: pd.DataFrame,
    tab_nas: pd.DataFrame,
    tab_diff: pd.DataFrame,
    nome_sheet: str = "Comparativo_CFOP",
    bloco_fonte: str = "C190",
    *,
    sheet_dados_cliente: str = "",
    sheet_dados_nascel: str = "",
    df_cli_export: pd.DataFrame | None = None,
    df_nas_export: pd.DataFrame | None = None,
) -> None:
    """
    Três blocos: T1/T2 com SUMIFS nas abas de detalhe após forçar CFOP como texto e VL_* como número.
    Tabela 3 continua em valores (Python).
    """
    sh = nome_sheet[:31]
    fmt_pink = writer.book.add_format(_FMT_XLSX_TITULO_BLOCO_PINK)
    fmt_laranja = writer.book.add_format(_FMT_XLSX_TITULO_BLOCO_LARANJA)
    fmt_txt_cfop = writer.book.add_format({"num_format": "@"})
    fmt_num_soma = writer.book.add_format({"num_format": "#,##0.00"})
    ncol1 = max(len(tab_cli.columns), 1)
    ncol2 = max(len(tab_nas.columns), 1)
    tab_diff_x = _comparativo_tabela3_para_excel(tab_diff)
    ncol3 = max(len(tab_diff_x.columns), 1)

    mapa_det = _COL_DETALHE_D190 if bloco_fonte == "D190" else _COL_DETALHE_C190

    start1 = 1
    tab_cli.to_excel(writer, sheet_name=sh, index=False, startrow=start1)
    ws = writer.sheets[sh]
    ws.merge_range(
        0,
        0,
        0,
        ncol1 - 1,
        f"Tabela 1 — Somatório por CFOP ({bloco_fonte}) — {ROTULO_CLIENTE}",
        fmt_pink,
    )
    _xlsx_reforcar_cfop_sped_coluna_a_texto(ws, tab_cli, start1, fmt_txt_cfop)

    lc = (
        _letras_colunas_soma(df_cli_export, mapa_det)
        if df_cli_export is not None
        else None
    )
    ln = (
        _letras_colunas_soma(df_nas_export, mapa_det)
        if df_nas_export is not None
        else None
    )
    usar_f = (
        bool(sheet_dados_cliente and sheet_dados_nascel and lc and ln)
        and df_cli_export is not None
        and df_nas_export is not None
        and len(df_cli_export) > 0
        and len(df_nas_export) > 0
    )
    if usar_f:
        ex_ini = 2
        ex_fim_cli = 1 + len(df_cli_export)
        ex_fim_nas = 1 + len(df_nas_export)
        _sobrescrever_somatorio_com_formulas(
            ws,
            tab_cli,
            start1,
            sheet_dados_cliente,
            lc,
            ex_ini,
            ex_fim_cli,
            fmt_num_soma,
        )

    if len(tab_cli) > 0:
        _xlsx_add_tabela_estilo_dinamica(
            ws,
            start1,
            start1 + len(tab_cli),
            0,
            ncol1 - 1,
            list(tab_cli.columns),
        )

    start2 = start1 + len(tab_cli) + 3
    tab_nas.to_excel(writer, sheet_name=sh, index=False, startrow=start2)
    ws.merge_range(
        start2 - 1,
        0,
        start2 - 1,
        ncol2 - 1,
        f"Tabela 2 — Somatório por CFOP ({bloco_fonte}) — {ROTULO_NASCEL}",
        fmt_laranja,
    )
    _xlsx_reforcar_cfop_sped_coluna_a_texto(ws, tab_nas, start2, fmt_txt_cfop)

    if usar_f:
        _sobrescrever_somatorio_com_formulas(
            ws,
            tab_nas,
            start2,
            sheet_dados_nascel,
            ln,
            ex_ini,
            ex_fim_nas,
            fmt_num_soma,
        )

    if len(tab_nas) > 0:
        _xlsx_add_tabela_estilo_dinamica(
            ws,
            start2,
            start2 + len(tab_nas),
            0,
            ncol2 - 1,
            list(tab_nas.columns),
        )

    start3 = start2 + len(tab_nas) + 3
    tab_diff_x.to_excel(writer, sheet_name=sh, index=False, startrow=start3)
    ws.merge_range(
        start3 - 1,
        0,
        start3 - 1,
        ncol3 - 1,
        f"Tabela 3 ({bloco_fonte}) — Diferenças (Cliente − Nascel), presença, divergência e observação",
        fmt_pink,
    )
    ws.set_column(0, min(25, ncol3), 16)
    obs_tit = "Observação (preencher à mão)"
    if obs_tit in tab_diff_x.columns:
        j_obs = list(tab_diff_x.columns).index(obs_tit)
        ws.set_column(j_obs, j_obs, 52)
    if len(tab_diff_x) > 0:
        _xlsx_add_tabela_estilo_dinamica(
            ws,
            start3,
            start3 + len(tab_diff_x),
            0,
            ncol3 - 1,
            list(tab_diff_x.columns),
        )


def _escrever_somatorio_cfop_unica(
    writer: pd.ExcelWriter,
    tab: pd.DataFrame,
    nome_sheet: str,
    bloco_fonte: str,
    rotulo_lado: str,
    sheet_dados: str,
    df_export: pd.DataFrame | None,
) -> None:
    """Uma só tabela tipo «Tabela 1» do comparativo, com SUMIFS no detalhe quando houver linhas."""
    sh = nome_sheet[:31]
    fmt = writer.book.add_format(_FMT_XLSX_TITULO_BLOCO)
    fmt_txt_cfop = writer.book.add_format({"num_format": "@"})
    fmt_num_soma = writer.book.add_format({"num_format": "#,##0.00"})
    ncol = max(len(tab.columns), 1)
    start = 1
    tab.to_excel(writer, sheet_name=sh, index=False, startrow=start)
    ws = writer.sheets[sh]
    ws.merge_range(
        0,
        0,
        0,
        ncol - 1,
        f"Tabela 1 — Somatório por CFOP ({bloco_fonte}) — {rotulo_lado}",
        fmt,
    )
    _xlsx_reforcar_cfop_sped_coluna_a_texto(ws, tab, start, fmt_txt_cfop)
    mapa_det = _COL_DETALHE_D190 if bloco_fonte == "D190" else _COL_DETALHE_C190
    letras = (
        _letras_colunas_soma(df_export, mapa_det) if df_export is not None else None
    )
    if (
        sheet_dados
        and letras
        and df_export is not None
        and len(df_export) > 0
    ):
        ex_ini = 2
        ex_fim = 1 + len(df_export)
        _sobrescrever_somatorio_com_formulas(
            ws,
            tab,
            start,
            sheet_dados,
            letras,
            ex_ini,
            ex_fim,
            fmt_num_soma,
        )
    if len(tab) > 0:
        _xlsx_add_tabela_estilo_dinamica(
            ws, start, start + len(tab), 0, ncol - 1, list(tab.columns)
        )
    ws.set_column(0, min(25, ncol), 16)


def _digits_chave(v) -> str:
    return re.sub(r"\D", "", _norm_val(v))


def _cnpj_emitente_de_chave_nfe_44(chv: str) -> str:
    """
    CNPJ do emitente na chave NF-e / NFC-e (44 dígitos): posições 7 a 20 (base 1),
    conforme layout da chave de acesso (Ajuste SINIEF 07/05).
    """
    ch = _digits_chave(chv)
    if len(ch) != 44:
        return ""
    return ch[6:20]


def _mapa_destaque_celulas_cfop_comparativo(
    df_cli: pd.DataFrame,
    df_nas: pd.DataFrame,
    fonte_cfop: str,
    tol: float = 0.02,
) -> dict[tuple[str, str], dict[str, str]]:
    """
    (chave44, CFOP) → { coluna_interna_VL_* ou «CFOP» : categoria_cor }.
    Todas as divergências usam a mesma categoria visual «divergencia».
    """
    pc = _agregar_por_chave_cfop(df_cli, fonte_cfop)
    pn = _agregar_por_chave_cfop(df_nas, fonte_cfop)
    if pc.empty and pn.empty:
        return {}
    keys = ["CHV_44", "CFOP SPED"]
    m = pc.merge(pn, on=keys, how="outer", suffixes=("_Cliente", "_Nascel"))
    for col in _COLS_C190_SOMA:
        cl, cn = f"{col}_Cliente", f"{col}_Nascel"
        if cl not in m.columns:
            m[cl] = 0.0
        if cn not in m.columns:
            m[cn] = 0.0
        m[cl] = m[cl].map(_to_float_celula_br)
        m[cn] = m[cn].map(_to_float_celula_br)

    out: dict[tuple[str, str], dict[str, str]] = {}
    for _, row in m.iterrows():
        ch_raw = str(row.get("CHV_44", "")).strip()
        ch = _digits_chave(ch_raw)
        if len(ch) < 10:
            continue
        cf = str(_norm_cfop_sped(row.get("CFOP SPED", "")))
        s_cli = sum(abs(float(row.get(f"{c}_Cliente", 0) or 0)) for c in _COLS_C190_SOMA)
        s_nas = sum(abs(float(row.get(f"{c}_Nascel", 0) or 0)) for c in _COLS_C190_SOMA)
        if s_nas < tol and s_cli >= tol:
            origem_linha = "só_Cliente"
        elif s_cli < tol and s_nas >= tol:
            origem_linha = "só_Nascel"
        else:
            origem_linha = "ambos"

        por_col: dict[str, str] = {}
        for metric_col in _COLS_C190_SOMA:
            cl, cn = f"{metric_col}_Cliente", f"{metric_col}_Nascel"
            vc = float(row.get(cl, 0) or 0)
            vn = float(row.get(cn, 0) or 0)
            if abs(round(vc - vn, 2)) < tol:
                continue
            if origem_linha in ("só_Cliente", "só_Nascel"):
                por_col[metric_col] = "divergencia"
            else:
                por_col[metric_col] = "divergencia"
        if por_col and origem_linha in ("só_Cliente", "só_Nascel"):
            por_col["CFOP"] = "divergencia"
        if por_col:
            out[(ch, cf)] = por_col

    # Mesma nota: CFOP escriturados diferentes entre SPED (ex.: um lado tem CFOP a mais) —
    # mesmo quando num CFOP comum os totais batem, tem de aparecer como erro na coluna CFOP.
    map_cli_cf = _mapa_cfops_por_ch_agregado(pc, tol)
    map_nas_cf = _mapa_cfops_por_ch_agregado(pn, tol)
    for ch in set(map_cli_cf) | set(map_nas_cf):
        sc = map_cli_cf.get(ch, set())
        sn = map_nas_cf.get(ch, set())
        if not sc or not sn or sc == sn:
            continue
        for cf in sc | sn:
            k = (ch, str(_norm_cfop_sped(cf)))
            cur = dict(out.get(k, {}))
            cur["CFOP"] = "divergencia"
            out[k] = cur

    return out


def _enriquecer_c190_com_produtos_c170(
    c190: pd.DataFrame,
    c100: pd.DataFrame,
    c170: pd.DataFrame | None,
    t0200: pd.DataFrame | None,
) -> pd.DataFrame:
    """
    Acrescenta COD_ITEM e DESCR_ITEM por linha C190, a partir dos itens C170 (mesma NF + CFOP;
    agrega todos os CST desse CFOP) e cadastro 0200 (DESCR_ITEM). Sem C170/0200, devolve vazio.
    """
    if c190 is None or c190.empty:
        return c190 if c190 is not None else pd.DataFrame()
    out = c190.copy()
    out["COD_ITEM"] = ""
    out["DESCR_ITEM"] = ""
    if c170 is None or c170.empty or "COD_ITEM" not in c170.columns:
        return out
    keys = [k for k in ("NUM_DOC", "SER", "COD_PART") if k in c100.columns and k in c170.columns]
    if len(keys) < 1 or c100.empty or "CHV_NFE" not in c100.columns:
        return out
    lkp = c100[keys + ["CHV_NFE"]].copy()
    lkp["CHV_NFE"] = lkp["CHV_NFE"].map(_digits_chave)
    lkp = lkp.drop_duplicates(subset=keys, keep="first")
    ci = c170.merge(lkp, on=keys, how="left")
    if "CHV_NFE" not in ci.columns:
        return out
    ci["_chv44"] = ci["CHV_NFE"].map(_digits_chave)
    if "CFOP" not in ci.columns or "CST_ICMS" not in ci.columns:
        return out
    ci["CFOP_n"] = ci["CFOP"].map(_norm_cfop_sped)
    if "DESCR_COMPL" not in ci.columns:
        ci["DESCR_COMPL"] = ""
    if t0200 is not None and not t0200.empty and "COD_ITEM" in t0200.columns and "DESCR_ITEM" in t0200.columns:
        t2 = t0200[["COD_ITEM", "DESCR_ITEM"]].drop_duplicates(subset=["COD_ITEM"], keep="last")
        ci = ci.merge(t2.rename(columns={"DESCR_ITEM": "_DESCR_0200"}), on="COD_ITEM", how="left")
    else:
        ci["_DESCR_0200"] = ""
    dc = ci["DESCR_COMPL"].fillna("").astype(str).str.strip()
    d0 = ci["_DESCR_0200"].fillna("").astype(str).str.strip()
    ci["_d_item"] = dc.where(dc != "", d0).str.slice(0, 4000)

    def _pack_grupo(g: pd.DataFrame) -> tuple[str, str]:
        pares: list[tuple[str, str]] = []
        vistos: set[tuple[str, str]] = set()
        for _, r in g.iterrows():
            cod = _norm_val(r.get("COD_ITEM", ""))
            des = _norm_val(r.get("_d_item", ""))
            key = (cod, des)
            if key in vistos:
                continue
            vistos.add(key)
            if cod or des:
                pares.append((cod, des))
        if not pares:
            return "", ""
        codes = "; ".join(a for a, _ in pares if a)[:8000]
        descs = "; ".join((b if b else "—") for _, b in pares)[:8000]
        return codes, descs

    lut: dict[tuple[str, str], tuple[str, str]] = {}
    for name, g in ci.groupby(["_chv44", "CFOP_n"], dropna=False, sort=False):
        ch, cf = name
        lut[(str(ch), str(cf))] = _pack_grupo(g)

    if "CHV_NFE" not in out.columns:
        out["CHV_NFE"] = ""
    out["_chv44"] = out["CHV_NFE"].map(_digits_chave)
    cods: list[str] = []
    descs: list[str] = []
    for _, row in out.iterrows():
        key = (str(row["_chv44"]), str(_norm_cfop_sped(row.get("CFOP", ""))))
        co, de = lut.get(key, ("", ""))
        cods.append(co)
        descs.append(de)
    out["COD_ITEM"] = cods
    out["DESCR_ITEM"] = descs
    return out.drop(columns=["_chv44"], errors="ignore")


def _xlsx_realcar_celulas_destaque_metrica(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df_excel: pd.DataFrame,
    titulo_chave: str,
    titulo_cfop: str,
    titulo_cst: str | None,
    mapa_celulas: dict[tuple[str, str], dict[str, str]],
    mapa_headers: dict[str, str],
    colunas_contexto_internos: tuple[str, ...],
) -> None:
    """
    Colorir colunas de contexto (cinza) e células com divergência (um único tom de realce).
    Cruzamento por chave+CFOP (sem CST).
    """
    if not mapa_celulas or df_excel is None or df_excel.empty:
        return
    if sheet_name not in writer.sheets:
        return
    for tit in (titulo_chave, titulo_cfop):
        if tit not in df_excel.columns:
            return
    if titulo_cst is not None and titulo_cst not in df_excel.columns:
        return
    jch = list(df_excel.columns).index(titulo_chave)
    jcf = list(df_excel.columns).index(titulo_cfop)
    ws = writer.sheets[sheet_name]
    wb = writer.book
    fmts_num: dict[str, object] = {}
    for cat, props in _DESTAQUE_XLSX_FMT_POR_CATEG.items():
        d = dict(props)
        d["num_format"] = "#,##0.00"
        fmts_num[cat] = wb.add_format(d)
    fmt_ctx = wb.add_format({**_DESTAQUE_XLSX_FMT_CONTEXTO, "num_format": "@"})
    fmts_txt_cfop: dict[str, object] = {}
    d0 = dict(_DESTAQUE_XLSX_FMT_POR_CATEG["divergencia"])
    d0["num_format"] = "@"
    fmts_txt_cfop["divergencia"] = wb.add_format(d0)

    for i in range(len(df_excel)):
        ch = _digits_chave(df_excel.iloc[i, jch])
        cf = str(_norm_cfop_sped(df_excel.iloc[i, jcf]))
        inner = mapa_celulas.get((ch, cf))
        if not inner:
            continue
        for interno in colunas_contexto_internos:
            j = _col_excel_idx(df_excel, interno, mapa_headers)
            if j is None:
                continue
            raw = df_excel.iloc[i, j]
            if interno == "CFOP":
                txt = "" if pd.isna(raw) else str(_norm_cfop_sped(raw))
            else:
                txt = "" if pd.isna(raw) else str(_norm_val(raw))
            ws.write_string(i + 1, j, txt, fmt_ctx)
        for metric_col, categoria in inner.items():
            if metric_col == "CFOP":
                ws.write_string(
                    i + 1,
                    jcf,
                    str(_norm_cfop_sped(df_excel.iloc[i, jcf])),
                    fmts_txt_cfop["divergencia"],
                )
                continue
            fmt = fmts_num.get(categoria) or fmts_num.get("divergencia")
            if fmt is None:
                continue
            j = _col_excel_idx(df_excel, metric_col, mapa_headers)
            if j is None:
                continue
            val = _to_float_celula_br(df_excel.iloc[i, j])
            ws.write_number(i + 1, j, val, fmt)


def _garantir_chv_em_c190(c100: pd.DataFrame, c190: pd.DataFrame) -> pd.DataFrame:
    """SPED .txt (spedlib): C190 não traz CHV_NFE nas colunas — cruza com C100 por NUM/SER/PART."""
    if c190.empty:
        return c190
    if "CHV_NFE" in c190.columns:
        s = c190["CHV_NFE"].map(_digits_chave)
        if s.astype(str).str.len().ge(40).any():
            return c190
    keys = [k for k in ("NUM_DOC", "SER", "COD_PART") if k in c100.columns and k in c190.columns]
    if len(keys) < 2 or c100.empty or "CHV_NFE" not in c100.columns:
        out = c190.copy()
        if "CHV_NFE" not in out.columns:
            out["CHV_NFE"] = ""
        return out
    lkp = c100[keys + ["CHV_NFE"]].copy()
    lkp["CHV_NFE"] = lkp["CHV_NFE"].map(_digits_chave)
    lkp = lkp.drop_duplicates(subset=keys, keep="first")
    base = c190.drop(columns=["CHV_NFE"], errors="ignore")
    extra = [c for c in ("IND_EMIT", "IND_OPER", "CNPJ") if c in c100.columns and c not in base.columns]
    if extra:
        add = c100[keys + extra].drop_duplicates(subset=keys, keep="first")
        lkp = lkp.merge(add, on=keys, how="left")
    return base.merge(lkp, on=keys, how="left")


def _garantir_chv_em_d190(d100: pd.DataFrame, d190: pd.DataFrame) -> pd.DataFrame:
    if d190.empty:
        return d190
    for c in ("CHV_DOC", "CHV_NFE"):
        if c in d190.columns:
            s = d190[c].map(_digits_chave)
            if s.astype(str).str.len().ge(40).any():
                return d190
            break
    keys = [k for k in ("NUM_DOC", "SER", "COD_PART") if k in d100.columns and k in d190.columns]
    ch_col = "CHV_DOC" if "CHV_DOC" in d100.columns else ("CHV_NFE" if "CHV_NFE" in d100.columns else None)
    if len(keys) < 2 or d100.empty or not ch_col:
        return d190
    lkp = d100[keys + [ch_col]].copy()
    lkp[ch_col] = lkp[ch_col].map(_digits_chave)
    lkp = lkp.rename(columns={ch_col: "CHV_DOC"})
    lkp = lkp.drop_duplicates(subset=keys, keep="first")
    base = d190.drop(columns=[c for c in ("CHV_DOC", "CHV_NFE") if c in d190.columns], errors="ignore")
    out = base.merge(lkp, on=keys, how="left")
    return out


def _serie_chv44(df: pd.DataFrame, fonte_cfop: str | None = None) -> pd.Series:
    for c in ("CHV_DOC", "CHV_NFE", "CHV_CTE"):
        if c in df.columns:
            return df[c].map(_digits_chave)
    return pd.Series([""] * len(df), index=df.index, dtype=object)


def _agregar_por_chave_cfop(df: pd.DataFrame, fonte_cfop: str) -> pd.DataFrame:
    """
    Soma bases e impostos por nota (44 dígitos) + CFOP, **sem** separar por CST —
    alinhado ao confronto «CFOP + valores» entre os dois SPED.
    """
    if df.empty:
        return pd.DataFrame(columns=["CHV_44", "NUM_DOC", "CFOP SPED"] + _COLS_C190_SOMA)
    g = df.copy()
    g["CHV_44"] = _serie_chv44(g, fonte_cfop)
    if "NUM_DOC" not in g.columns:
        g["NUM_DOC"] = ""
    g["NUM_DOC"] = g["NUM_DOC"].map(_norm_val)
    g["CFOP SPED"] = g["CFOP"].map(_norm_cfop_sped) if "CFOP" in g.columns else "(vazio)"
    for c in _COLS_C190_SOMA:
        if c not in g.columns:
            g[c] = 0.0
        else:
            g[c] = g[c].map(_to_float_celula_br)
    agg = {c: "sum" for c in _COLS_C190_SOMA}
    agg["NUM_DOC"] = "first"
    return g.groupby(["CHV_44", "CFOP SPED"], dropna=False).agg(agg).reset_index()


def _mapa_cfops_por_ch_agregado(p: pd.DataFrame, tol: float = 0.02) -> dict[str, set[str]]:
    """chave44 → conjunto de CFOP com soma ≥ tol (agregado por chave+CFOP, sem CST)."""
    if p is None or p.empty:
        return {}
    out: dict[str, set[str]] = defaultdict(set)
    for _, row in p.iterrows():
        ch = _digits_chave(str(row.get("CHV_44", "")).strip())
        if len(ch) < 10:
            continue
        cf = str(_norm_cfop_sped(row.get("CFOP SPED", "")))
        s = sum(abs(float(row.get(c, 0) or 0)) for c in _COLS_C190_SOMA)
        if s >= tol:
            out[ch].add(cf)
    return dict(out)


def _montar_onde_agir_cfop_por_nota(
    df_cli: pd.DataFrame,
    df_nas: pd.DataFrame,
    fonte_cfop: str,
    map_status_cli: dict[str, str],
    map_status_nas: dict[str, str],
    max_linhas: int = 500_000,
) -> pd.DataFrame:
    """
    Uma linha por nota (chave 44): coluna «O que» com etiquetas curtas (CFOP diverge, ICMS diverge, …).
    """
    pc = _agregar_por_chave_cfop(df_cli, fonte_cfop)
    pn = _agregar_por_chave_cfop(df_nas, fonte_cfop)
    if pc.empty and pn.empty:
        return pd.DataFrame()

    keys = ["CHV_44", "CFOP SPED"]
    m = pc.merge(pn, on=keys, how="outer", suffixes=("_Cliente", "_Nascel"))
    for col in _COLS_C190_SOMA:
        cl, cn = f"{col}_Cliente", f"{col}_Nascel"
        if cl not in m.columns:
            m[cl] = 0.0
        if cn not in m.columns:
            m[cn] = 0.0
        m[cl] = m[cl].map(_to_float_celula_br)
        m[cn] = m[cn].map(_to_float_celula_br)

    ndc = "NUM_DOC_Cliente" if "NUM_DOC_Cliente" in m.columns else None
    ndn = "NUM_DOC_Nascel" if "NUM_DOC_Nascel" in m.columns else None
    tol = 0.02
    linhas: list[dict] = []
    truncou = False
    map_cli_cf = _mapa_cfops_por_ch_agregado(pc, tol)
    map_nas_cf = _mapa_cfops_por_ch_agregado(pn, tol)

    m["_ch_key"] = m["CHV_44"].map(lambda x: _digits_chave(str(x).strip()))
    m = m[m["_ch_key"].str.len() >= 10]

    for ch_key, grupo in m.groupby("_ch_key", sort=False):
        if len(linhas) >= max_linhas:
            truncou = True
            break
        tags: set[str] = set()
        set_cli_cf = map_cli_cf.get(ch_key, set())
        set_nas_cf = map_nas_cf.get(ch_key, set())
        if set_cli_cf and set_nas_cf and set_cli_cf != set_nas_cf:
            tags.add(_TAG_ONDE_CFOP_DIVERGE)

        num_doc = ""
        chv_raw = ""
        for _, row in grupo.iterrows():
            chv = str(row["CHV_44"]).strip()
            if not chv_raw and chv:
                chv_raw = chv
            if not num_doc and ndc and pd.notna(row.get(ndc)):
                num_doc = str(row[ndc]).strip()
            if not num_doc and ndn and pd.notna(row.get(ndn)):
                num_doc = str(row[ndn]).strip()

            s_cli = sum(abs(float(row[f"{c}_Cliente"])) for c in _COLS_C190_SOMA)
            s_nas = sum(abs(float(row[f"{c}_Nascel"])) for c in _COLS_C190_SOMA)
            um_lado_zero_no_triplo = (s_nas < tol and s_cli >= tol) or (s_cli < tol and s_nas >= tol)
            if (
                um_lado_zero_no_triplo
                and set_cli_cf
                and set_nas_cf
                and set_cli_cf != set_nas_cf
            ):
                origem_linha = "cfop_diverge"
            elif s_nas < tol and s_cli >= tol:
                origem_linha = "só_Cliente"
            elif s_cli < tol and s_nas >= tol:
                origem_linha = "só_Nascel"
            else:
                origem_linha = "ambos"

            for col in _COLS_C190_SOMA:
                vc = float(row[f"{col}_Cliente"])
                vn = float(row[f"{col}_Nascel"])
                if abs(round(vc - vn, 2)) < tol:
                    continue
                tags.add(_TAG_ONDE_METRICA.get(col, f"{col} diverge"))
                if origem_linha == "só_Cliente":
                    tags.add(_TAG_ONDE_SO_CLIENTE)
                elif origem_linha == "só_Nascel":
                    tags.add(_TAG_ONDE_SO_NASCEL)

        if not tags:
            continue

        nota_ref = _onde_agir_ref_nota(num_doc, chv_raw or ch_key)
        chave44 = _digits_chave(str(chv_raw or ch_key or "").strip())
        linhas.append(
            {
                "#": len(linhas) + 1,
                "Nota": nota_ref,
                "Chave_acesso": chave44 if len(chave44) == 44 else "—",
                "CFOP_CLIENTE": _formatar_conjunto_cfop_excel(set_cli_cf),
                "CFOP_NASCEL": _formatar_conjunto_cfop_excel(set_nas_cf),
                "Status": _status_onde_agir_linha(ch_key, map_status_cli, map_status_nas),
                "O_que": _onde_agir_tags_ordenadas(tags),
            }
        )

    if truncou:
        linhas.append(
            {
                "#": len(linhas) + 1,
                "Nota": "—",
                "Chave_acesso": "—",
                "CFOP_CLIENTE": "—",
                "CFOP_NASCEL": "—",
                "Status": "—",
                "O_que": f"Mais notas que o limite ({max_linhas}). Reduza o período.",
            }
        )
    return pd.DataFrame(linhas)


def _montar_onde_agir_cfop_somente_agregado(comp: pd.DataFrame, bloco_fonte: str) -> pd.DataFrame:
    div = comp[comp["Divergente"] == "SIM"]
    if div.empty:
        return pd.DataFrame(
            [
                {
                    "#": 1,
                    "Nota": "—",
                    "Chave_acesso": "—",
                    "Status": "—",
                    "O_que": "Sem divergência (por CFOP).",
                }
            ]
        )
    linhas = []
    for i, (_, row) in enumerate(div.iterrows(), start=1):
        cf = row["CFOP SPED"]
        pres = row.get("Presenca_CFOP", "Ambos")
        if pres != "Ambos":
            oque = "Só num SPED"
        else:
            oque = "Totais divergem"
        linhas.append(
            {"#": i, "Nota": str(cf), "Chave_acesso": "—", "Status": "—", "O_que": oque}
        )
    return pd.DataFrame(linhas)


def montar_onde_agir_cfop(
    comp: pd.DataFrame,
    bloco_fonte: str = "C190",
    df_cli: pd.DataFrame | None = None,
    df_nas: pd.DataFrame | None = None,
    df_head_cli: pd.DataFrame | None = None,
    df_head_nas: pd.DataFrame | None = None,
) -> pd.DataFrame:
    if comp.empty:
        hint = (
            "Sem D190: carregar EFD (.txt) com D100+D190 ou Excel com essas abas."
            if bloco_fonte == "D190"
            else "Sem C190: carregar EFD (.txt) com C100+C190 ou Excel com essas abas."
        )
        return pd.DataFrame(
            [
                {
                    "#": 1,
                    "Nota": "—",
                    "Chave_acesso": "—",
                    "Status": "—",
                    "O_que": hint,
                }
            ]
        )

    if df_cli is not None and df_nas is not None and (not df_cli.empty or not df_nas.empty):
        map_sc = _mapa_status_por_chave_doc(df_head_cli)
        map_sn = _mapa_status_por_chave_doc(df_head_nas)
        det = _montar_onde_agir_cfop_por_nota(
            df_cli, df_nas, bloco_fonte, map_sc, map_sn
        )
        if not det.empty:
            return det

    return _montar_onde_agir_cfop_somente_agregado(comp, bloco_fonte)


def _df_com_coluna_lado(df: pd.DataFrame, lado: str) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame(columns=["LADO"])
    if df.empty:
        return pd.DataFrame(columns=["LADO"] + list(df.columns))
    out = df.copy()
    out.insert(0, "LADO", lado)
    return out


def _df_export_sped_com_cfop_soma(df: pd.DataFrame, lado: str) -> pd.DataFrame:
    """Inclui CFOP_SPED_soma (texto) para o critério do SUMIFS coincidir com a coluna A do comparativo."""
    x = _df_com_coluna_lado(df, lado).copy()
    if x.empty:
        x[_COL_CFOP_SOM] = pd.Series(dtype=object)
        return x
    if "CFOP" in x.columns:
        x[_COL_CFOP_SOM] = x["CFOP"].map(_norm_cfop_sped).astype(str)
    else:
        x[_COL_CFOP_SOM] = "(vazio)"
    return x


def _ler_bytes_upload_sped(file_obj) -> bytes:
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    return file_obj.read()


def _carregar_blocos_de_bytes(data: bytes, filename: str) -> dict[str, pd.DataFrame]:
    return carregar_blocos_sped_completos(io.BytesIO(data), filename)


def _carregar_dois_speds_em_paralelo(
    file_sped_cliente,
    file_sped_nascel,
    nome_arquivo_cliente: str,
    nome_arquivo_nascel: str,
) -> tuple[dict[str, pd.DataFrame], dict[str, pd.DataFrame]]:
    """
    Copia cada upload para memória e faz o parse dos dois SPED em paralelo (dois ficheiros ao mesmo tempo).
    Isto é só desempenho: cada SPED continua independente — não funde nem soma os dois num único conjunto de dados.
    """
    dc = _ler_bytes_upload_sped(file_sped_cliente)
    dn = _ler_bytes_upload_sped(file_sped_nascel)
    with ThreadPoolExecutor(max_workers=2) as ex:
        fut_c = ex.submit(_carregar_blocos_de_bytes, dc, nome_arquivo_cliente)
        fut_n = ex.submit(_carregar_blocos_de_bytes, dn, nome_arquivo_nascel)
        bl_c = fut_c.result()
        bl_n = fut_n.result()
    return bl_c, bl_n


def _escrever_um_bloco_comparativo_cfop(
    writer: pd.ExcelWriter,
    bl_c: dict[str, pd.DataFrame],
    bl_n: dict[str, pd.DataFrame],
    fonte_cfop: str,
    onde_sheet: str,
) -> bool:
    """
    Escreve no mesmo workbook um bloco completo: detalhe (C190 NF-e ou D190 CT-e),
    folha Comparativo_CFOP_* e folha «onde agir» com nome `onde_sheet`.
    Devolve False se não houver linhas C190/D190 em nenhum dos dois SPED (bloco ignorado).
    """
    if fonte_cfop == "C190":
        bl_c["C190"] = _garantir_chv_em_c190(bl_c["C100"], bl_c["C190"])
        bl_n["C190"] = _garantir_chv_em_c190(bl_n["C100"], bl_n["C190"])
    else:
        bl_c["D190"] = _garantir_chv_em_d190(bl_c["D100"], bl_c["D190"])
        bl_n["D190"] = _garantir_chv_em_d190(bl_n["D100"], bl_n["D190"])

    chave_df = "D190" if fonte_cfop == "D190" else "C190"
    df_c = bl_c[chave_df]
    df_n = bl_n[chave_df]

    if df_c.empty and df_n.empty:
        return False

    tab_c = tabela_somatorio_cfop_estilo_excel(df_c)
    tab_n = tabela_somatorio_cfop_estilo_excel(df_n)
    agg_c = _agg_cfop_sem_total(df_c)
    agg_n = _agg_cfop_sem_total(df_n)
    comp = _comparativo_cfop_merge(agg_c, agg_n)
    if fonte_cfop == "C190":
        _hc, _hn = bl_c.get("C100"), bl_n.get("C100")
    else:
        _hc, _hn = bl_c.get("D100"), bl_n.get("D100")
    head_c = _hc if isinstance(_hc, pd.DataFrame) else pd.DataFrame()
    head_n = _hn if isinstance(_hn, pd.DataFrame) else pd.DataFrame()
    onde = montar_onde_agir_cfop(
        comp,
        bloco_fonte=fonte_cfop,
        df_cli=df_c,
        df_nas=df_n,
        df_head_cli=head_c,
        df_head_nas=head_n,
    )
    mapa_destaque_celulas = _mapa_destaque_celulas_cfop_comparativo(df_c, df_n, fonte_cfop)

    d100_cli = _df_com_coluna_lado(bl_c["D100"], ROTULO_CLIENTE)
    d100_nas = _df_com_coluna_lado(bl_n["D100"], ROTULO_NASCEL)
    d100_cli_pl = _df_c100_planilha(d100_cli)
    d100_nas_pl = _df_c100_planilha(d100_nas)
    d190_cli = _df_export_d190_planilha(bl_c["D190"], ROTULO_CLIENTE)
    d190_nas = _df_export_d190_planilha(bl_n["D190"], ROTULO_NASCEL)

    nome_comp = "Comparativo_CFOP_C190" if fonte_cfop == "C190" else "Comparativo_CFOP_D190"

    if fonte_cfop == "C190":
        sh_c190_cli, sh_c190_nas = "Cliente", "Nascel"
        c170_c = bl_c.get("C170") if isinstance(bl_c.get("C170"), pd.DataFrame) else pd.DataFrame()
        c170_n = bl_n.get("C170") if isinstance(bl_n.get("C170"), pd.DataFrame) else pd.DataFrame()
        t0200_c = bl_c.get("0200") if isinstance(bl_c.get("0200"), pd.DataFrame) else pd.DataFrame()
        t0200_n = bl_n.get("0200") if isinstance(bl_n.get("0200"), pd.DataFrame) else pd.DataFrame()
        c190_c_en = _enriquecer_c190_com_produtos_c170(
            bl_c["C190"].copy(), bl_c["C100"], c170_c, t0200_c
        )
        c190_n_en = _enriquecer_c190_com_produtos_c170(
            bl_n["C190"].copy(), bl_n["C100"], c170_n, t0200_n
        )
        c190_cli = _df_export_c190_planilha(c190_c_en, ROTULO_CLIENTE)
        c190_nas = _df_export_c190_planilha(c190_n_en, ROTULO_NASCEL)
        c190_cli.to_excel(writer, index=False, sheet_name=sh_c190_cli)
        c190_nas.to_excel(writer, index=False, sheet_name=sh_c190_nas)
        tit_chv = _COL_DETALHE_C190["CHV_NFE"]
        tit_cf = _COL_DETALHE_C190["CFOP"]
        _xlsx_reforcar_detalhe_texto_e_numero(
            writer, sh_c190_cli, c190_cli, _COL_DETALHE_C190
        )
        _xlsx_reforcar_detalhe_texto_e_numero(
            writer, sh_c190_nas, c190_nas, _COL_DETALHE_C190
        )
        if len(c190_cli) > 0:
            _xlsx_tabela_sobre_df_escrito(writer.sheets[sh_c190_cli], c190_cli)
        if len(c190_nas) > 0:
            _xlsx_tabela_sobre_df_escrito(writer.sheets[sh_c190_nas], c190_nas)
        _xlsx_realcar_celulas_destaque_metrica(
            writer,
            sh_c190_cli,
            c190_cli,
            tit_chv,
            tit_cf,
            None,
            mapa_destaque_celulas,
            _COL_DETALHE_C190,
            _COL_CTX_DESTAQUE_C190,
        )
        _xlsx_realcar_celulas_destaque_metrica(
            writer,
            sh_c190_nas,
            c190_nas,
            tit_chv,
            tit_cf,
            None,
            mapa_destaque_celulas,
            _COL_DETALHE_C190,
            _COL_CTX_DESTAQUE_C190,
        )
        abas_largura = (sh_c190_cli, sh_c190_nas)
        sheet_dados_c, sheet_dados_n = sh_c190_cli, sh_c190_nas
    else:
        c190_cli = _df_export_c190_planilha(bl_c["C190"], ROTULO_CLIENTE)
        c190_nas = _df_export_c190_planilha(bl_n["C190"], ROTULO_NASCEL)
        d100_cli_pl.to_excel(writer, index=False, sheet_name="D100_Cliente")
        d100_nas_pl.to_excel(writer, index=False, sheet_name="D100_Nascel")
        d190_cli.to_excel(writer, index=False, sheet_name="D190_Cliente")
        d190_nas.to_excel(writer, index=False, sheet_name="D190_Nascel")
        tit_cte = _COL_DETALHE_D190["CHV_DOC"]
        tit_cf_d = _COL_DETALHE_D190["CFOP"]
        _xlsx_reforcar_detalhe_texto_e_numero(
            writer, "D190_Cliente", d190_cli, _COL_DETALHE_D190
        )
        _xlsx_reforcar_detalhe_texto_e_numero(
            writer, "D190_Nascel", d190_nas, _COL_DETALHE_D190
        )
        for nm, dff in (
            ("D100_Cliente", d100_cli_pl),
            ("D100_Nascel", d100_nas_pl),
            ("D190_Cliente", d190_cli),
            ("D190_Nascel", d190_nas),
        ):
            if len(dff) > 0:
                _xlsx_tabela_sobre_df_escrito(writer.sheets[nm], dff)
        _xlsx_realcar_celulas_destaque_metrica(
            writer,
            "D190_Cliente",
            d190_cli,
            tit_cte,
            tit_cf_d,
            None,
            mapa_destaque_celulas,
            _COL_DETALHE_D190,
            _COL_CTX_DESTAQUE_D190,
        )
        _xlsx_realcar_celulas_destaque_metrica(
            writer,
            "D190_Nascel",
            d190_nas,
            tit_cte,
            tit_cf_d,
            None,
            mapa_destaque_celulas,
            _COL_DETALHE_D190,
            _COL_CTX_DESTAQUE_D190,
        )
        abas_largura = (
            "D100_Cliente",
            "D100_Nascel",
            "D190_Cliente",
            "D190_Nascel",
        )
        sheet_dados_c, sheet_dados_n = "D190_Cliente", "D190_Nascel"

    _escrever_comparativo_cfop_tres_tabelas(
        writer,
        tab_c,
        tab_n,
        comp,
        nome_sheet=nome_comp,
        bloco_fonte=fonte_cfop,
        sheet_dados_cliente=sheet_dados_c,
        sheet_dados_nascel=sheet_dados_n,
        df_cli_export=c190_cli if fonte_cfop == "C190" else d190_cli,
        df_nas_export=c190_nas if fonte_cfop == "C190" else d190_nas,
    )

    onde_pl = _onde_agir_cfop_para_excel(onde)
    sh_onde = onde_sheet[:31]
    onde_pl.to_excel(writer, index=False, sheet_name=sh_onde)

    for name in abas_largura:
        if name in writer.sheets:
            writer.sheets[name].set_column(0, 40, 14)
    if fonte_cfop == "C190":
        tit_cod = _COL_DETALHE_C190["COD_ITEM"]
        tit_dsc = _COL_DETALHE_C190["DESCR_ITEM"]
        tit_emit = _COL_DETALHE_C190["CNPJ_EMIT_NFE"]
        for shx, dfx in (("Cliente", c190_cli), ("Nascel", c190_nas)):
            if shx not in writer.sheets or dfx is None or dfx.empty:
                continue
            wloc = writer.sheets[shx]
            for j, h in enumerate(dfx.columns):
                if h == tit_cod:
                    wloc.set_column(j, j, 16)
                elif h == tit_dsc:
                    wloc.set_column(j, j, 48)
                elif h == tit_emit:
                    wloc.set_column(j, j, 18)
        _xlsx_largura_coluna_chave(writer, "Cliente", c190_cli, _COL_DETALHE_C190)
        _xlsx_largura_coluna_chave(writer, "Nascel", c190_nas, _COL_DETALHE_C190)
    else:
        for sn, dff, mp in (
            ("D100_Cliente", d100_cli_pl, _COL_C100_PLAN),
            ("D100_Nascel", d100_nas_pl, _COL_C100_PLAN),
            ("D190_Cliente", d190_cli, _COL_DETALHE_D190),
            ("D190_Nascel", d190_nas, _COL_DETALHE_D190),
        ):
            _xlsx_largura_coluna_chave(writer, sn, dff, mp)
    if sh_onde in writer.sheets:
        w = writer.sheets[sh_onde]
        ncols = len(onde_pl.columns)
        for j in range(ncols):
            hdr = str(onde_pl.columns[j])
            if hdr.strip().lower() == "resumo" or "O que" in hdr or "diferente" in hdr.lower():
                w.set_column(j, j, 78)
            elif "status" in hdr.lower():
                w.set_column(j, j, 36)
            elif hdr.strip().lower().startswith("nota"):
                w.set_column(j, j, 22)
            elif "chave" in hdr.lower():
                w.set_column(j, j, 48)
            else:
                w.set_column(j, j, 6)
    _xlsx_escrever_legenda_destaque_cfop(writer)
    return True


def _gerar_excel_cfop_comparativo_de_blocos(
    bl_c: dict[str, pd.DataFrame],
    bl_n: dict[str, pd.DataFrame],
    fonte_cfop: str,
):
    """
    Gera o Excel de comparativo CFOP a partir de blocos já carregados.
    Ordem final das abas: Comparativo_CFOP_* → Legenda_cores → ONDE_AGIR → detalhe.
    Altera bl_c/bl_n no ramo correspondente (CHV em C190 ou D190).
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        ok = _escrever_um_bloco_comparativo_cfop(
            writer, bl_c, bl_n, fonte_cfop, "ONDE_AGIR"
        )
        if not ok:
            if fonte_cfop == "D190":
                return None, (
                    "Não encontrei registros D190. No .txt, cada D190 deve vir após o D100 (CT-e); "
                    "no Excel, use abas D100 e D190."
                )
            return None, (
                "Não encontrei registros C190. No .txt, o C190 deve vir logo após o C100; "
                "no Excel, use abas C100 e C190 com colunas padrão EFD."
            )
        _xlsx_escrever_blocos_apur_difal_st(writer, bl_c, bl_n)
        _xlsx_aplicar_abas_rosa(writer)
    output.seek(0)
    ordem = (
        _ORDEM_ABAS_COMP_D190 if fonte_cfop == "D190" else _ORDEM_ABAS_COMP_C190
    )
    output = _xlsx_reordenar_abas_por_lista(output, ordem)
    return output, "Sucesso"


def _gerar_excel_cfop_comparativo_nfe_cte_de_blocos(
    bl_c: dict[str, pd.DataFrame],
    bl_n: dict[str, pd.DataFrame],
):
    """
    Um único .xlsx: comparativo por CFOP para **NF-e (C190)** e para **CT-e (D190)**.
    Ordem das abas: Comparativo_CFOP (C190 e D190) → Legenda_cores → Onde agir → detalhes.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        ok_nfe = _escrever_um_bloco_comparativo_cfop(
            writer, bl_c, bl_n, "C190", "ONDE_AGIR_NF_e"
        )
        ok_cte = _escrever_um_bloco_comparativo_cfop(
            writer, bl_c, bl_n, "D190", "ONDE_AGIR_CTe"
        )
        if not ok_nfe and not ok_cte:
            return None, (
                "Não encontrei registros C190 (NF-e) nem D190 (CT-e) nos dois SPED. "
                "No .txt, o C190 deve seguir o C100 e o D190 o D100; "
                "no Excel, use abas C100/C190 e D100/D190."
            )
        _xlsx_escrever_blocos_apur_difal_st(writer, bl_c, bl_n)
        _xlsx_aplicar_abas_rosa(writer)
    output.seek(0)
    output = _xlsx_reordenar_abas_por_lista(output, _ORDEM_ABAS_COMP_NFE_CTE)
    return output, "Sucesso"


def gerar_excel_cfop_comparativo_dois_speds(
    file_sped_cliente,
    file_sped_nascel,
    nome_arquivo_cliente: str = "",
    nome_arquivo_nascel: str = "",
    fonte_cfop: str = "C190",
):
    """
    Formato tipo relatório comparativo (duas origens): abas separadas para SPED Cliente e SPED Nascel,
    mais tabelas de confronto por CFOP (diferença Cliente−Nascel), não um único ficheiro “somado”.

    Abas conforme fonte_cfop:
    - C190: bloco C (C100, C190 × Cliente/Nascel) + comparativo + ONDE_AGIR.
    - D190: bloco D (D100, D190 × Cliente/Nascel) + comparativo + ONDE_AGIR.

    Os dois SPED são lidos em paralelo (ThreadPoolExecutor) — apenas para acelerar; cada um mantém a sua identidade.

    Para dois ficheiros .xlsx num ZIP com uma única dupla de leituras, use `gerar_zip_cfop_c190_e_d190_dois_speds`.
    """
    if fonte_cfop not in ("C190", "D190"):
        return None, "fonte_cfop deve ser C190 ou D190."

    try:
        bl_c, bl_n = _carregar_dois_speds_em_paralelo(
            file_sped_cliente,
            file_sped_nascel,
            nome_arquivo_cliente,
            nome_arquivo_nascel,
        )
        return _gerar_excel_cfop_comparativo_de_blocos(bl_c, bl_n, fonte_cfop)
    except Exception as e:
        return None, f"Erro: {str(e)}"


def gerar_excel_cfop_comparativo_nfe_e_cte_dois_speds(
    file_sped_cliente,
    file_sped_nascel,
    nome_arquivo_cliente: str = "",
    nome_arquivo_nascel: str = "",
):
    """
    Dois SPED em paralelo: **um** Excel com comparativo **NF-e (C190)** e **CT-e (D190)** —
    abas Cliente/Nascel + Comparativo_CFOP_C190 + ONDE_AGIR_NF_e, e bloco D100/D190 +
    Comparativo_CFOP_D190 + ONDE_AGIR_CTe (cada bloco só entra se houver linhas).
    """
    try:
        bl_c, bl_n = _carregar_dois_speds_em_paralelo(
            file_sped_cliente,
            file_sped_nascel,
            nome_arquivo_cliente,
            nome_arquivo_nascel,
        )
        return _gerar_excel_cfop_comparativo_nfe_cte_de_blocos(bl_c, bl_n)
    except Exception as e:
        return None, f"Erro: {str(e)}"


def gerar_zip_cfop_c190_e_d190_dois_speds(
    file_sped_cliente,
    file_sped_nascel,
    nome_arquivo_cliente: str = "",
    nome_arquivo_nascel: str = "",
):
    """
    Mesmos dois ficheiros Excel que gerar C190 e D190 à parte, mas com uma única leitura
    paralela Cliente + Nascel (evita parse duplicado dos .txt).
    """
    try:
        bl_c, bl_n = _carregar_dois_speds_em_paralelo(
            file_sped_cliente,
            file_sped_nascel,
            nome_arquivo_cliente,
            nome_arquivo_nascel,
        )
        bio_c, msg_c = _gerar_excel_cfop_comparativo_de_blocos(
            bl_c, bl_n, "C190"
        )
        bio_d, msg_d = _gerar_excel_cfop_comparativo_de_blocos(
            bl_c, bl_n, "D190"
        )
        if bio_c and bio_d:
            zbuf = io.BytesIO()
            with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr(
                    "Detetive_CFOP_C190_2SPED.xlsx",
                    bio_c.getvalue(),
                )
                zf.writestr(
                    "Detetive_CFOP_D190_2SPED.xlsx",
                    bio_d.getvalue(),
                )
            zbuf.seek(0)
            return zbuf, "Sucesso"
        err = " ".join(
            x
            for x in (
                (msg_c or "") if not bio_c else "",
                (msg_d or "") if not bio_d else "",
            )
            if x
        )
        return None, err or "Não foi possível gerar um ou ambos os relatórios."
    except Exception as e:
        return None, f"Erro: {str(e)}"


def gerar_excel_cfop_um_sped(
    file_sped,
    nome_arquivo: str = "",
    fonte_cfop: str = "C190",
    rotulo_lado: str = "",
):
    """
    Um único arquivo SPED: mesma «tabelinha» somatório por CFOP (estilo comparativo Tabela 1)
    e abas C100/C190 ou D100/D190 com coluna LADO. Sem Tabela 2/3 e sem ONDE_AGIR.
    """
    if fonte_cfop not in ("C190", "D190"):
        return None, "fonte_cfop deve ser C190 ou D190."
    if not rotulo_lado:
        rotulo_lado = ROTULO_CLIENTE
    try:
        if hasattr(file_sped, "seek"):
            file_sped.seek(0)
        bl = carregar_blocos_sped_completos(file_sped, nome_arquivo)
        if fonte_cfop == "C190":
            bl["C190"] = _garantir_chv_em_c190(bl["C100"], bl["C190"])
        else:
            bl["D190"] = _garantir_chv_em_d190(bl["D100"], bl["D190"])

        chave_df = "D190" if fonte_cfop == "D190" else "C190"
        df = bl[chave_df]

        if df.empty:
            if fonte_cfop == "D190":
                return None, (
                    "Não encontrei registros D190. No .txt, cada D190 deve vir após o D100 (CT-e); "
                    "no Excel, use abas D100 e D190."
                )
            return None, (
                "Não encontrei registros C190. No .txt, o C190 deve vir logo após o C100; "
                "no Excel, use abas C100 e C190 com colunas padrão EFD."
            )

        tab = tabela_somatorio_cfop_estilo_excel(df)
        c100_x = _df_c100_planilha(_df_com_coluna_lado(bl["C100"], rotulo_lado))
        d100_x = _df_c100_planilha(_df_com_coluna_lado(bl["D100"], rotulo_lado))
        c190_x = _df_export_c190_planilha(bl["C190"], rotulo_lado)
        d190_x = _df_export_d190_planilha(bl["D190"], rotulo_lado)

        nome_comp = (
            "Somatorio_CFOP_C190_um_SPED"
            if fonte_cfop == "C190"
            else "Somatorio_CFOP_D190_um_SPED"
        )
        sh_c100, sh_c190 = "C100", "C190"
        sh_d100, sh_d190 = "D100", "D190"

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            if fonte_cfop == "C190":
                c100_x.to_excel(writer, index=False, sheet_name=sh_c100)
                c190_x.to_excel(writer, index=False, sheet_name=sh_c190)
                _xlsx_reforcar_detalhe_texto_e_numero(
                    writer, sh_c190, c190_x, _COL_DETALHE_C190
                )
                abas_largura = (sh_c100, sh_c190)
                df_det = c190_x
                sh_det = sh_c190
            else:
                d100_x.to_excel(writer, index=False, sheet_name=sh_d100)
                d190_x.to_excel(writer, index=False, sheet_name=sh_d190)
                _xlsx_reforcar_detalhe_texto_e_numero(
                    writer, sh_d190, d190_x, _COL_DETALHE_D190
                )
                abas_largura = (sh_d100, sh_d190)
                df_det = d190_x
                sh_det = sh_d190

            _escrever_somatorio_cfop_unica(
                writer,
                tab,
                nome_comp,
                fonte_cfop,
                rotulo_lado,
                sh_det,
                df_det,
            )

            _xlsx_escrever_blocos_apur_difal_st(writer, bl, None)

            for name in abas_largura:
                if name in writer.sheets:
                    writer.sheets[name].set_column(0, 40, 14)
            if fonte_cfop == "C190":
                _xlsx_largura_coluna_chave(writer, sh_c100, c100_x, _COL_C100_PLAN)
                _xlsx_largura_coluna_chave(writer, sh_c190, c190_x, _COL_DETALHE_C190)
                _xlsx_largura_coluna_cnpj_emitente_nfe(writer, sh_c190, c190_x)
            else:
                _xlsx_largura_coluna_chave(writer, sh_d100, d100_x, _COL_C100_PLAN)
                _xlsx_largura_coluna_chave(writer, sh_d190, d190_x, _COL_DETALHE_D190)

            _xlsx_aplicar_abas_rosa(writer)
        output.seek(0)
        output = _xlsx_workbook_posprocesso_liberar(output)
        return output, "Sucesso"
    except Exception as e:
        return None, f"Erro: {str(e)}"


def gerar_excel_cfop_um_sped_completo(
    file_sped,
    nome_arquivo: str = "",
    rotulo_lado: str = "",
):
    """
    Um único SPED: abas C100/C190 + somatório CFOP (C190) e D100/D190 + somatório CFOP (D190)
    no mesmo Excel, para cada bloco que tiver linhas. Sem comparativo entre dois ficheiros.
    """
    if not rotulo_lado:
        rotulo_lado = ROTULO_CLIENTE
    try:
        if hasattr(file_sped, "seek"):
            file_sped.seek(0)
        bl = carregar_blocos_sped_completos(file_sped, nome_arquivo)
        bl["C190"] = _garantir_chv_em_c190(bl["C100"], bl["C190"])
        bl["D190"] = _garantir_chv_em_d190(bl["D100"], bl["D190"])

        tem_c = not bl["C190"].empty
        tem_d = not bl["D190"].empty
        if not tem_c and not tem_d:
            return None, (
                "Não encontrei registros C190 nem D190. No .txt, o C190 deve seguir o C100 e o D190 o D100; "
                "no Excel, use abas C100/C190 e D100/D190."
            )

        c100_x = _df_c100_planilha(_df_com_coluna_lado(bl["C100"], rotulo_lado))
        d100_x = _df_c100_planilha(_df_com_coluna_lado(bl["D100"], rotulo_lado))
        c190_x = _df_export_c190_planilha(bl["C190"], rotulo_lado)
        d190_x = _df_export_d190_planilha(bl["D190"], rotulo_lado)

        sh_c100, sh_c190 = "C100", "C190"
        sh_d100, sh_d190 = "D100", "D190"
        nome_comp_c = "Somatorio_CFOP_C190_um_SPED"
        nome_comp_d = "Somatorio_CFOP_D190_um_SPED"

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            if tem_c:
                tab_c = tabela_somatorio_cfop_estilo_excel(bl["C190"])
                c100_x.to_excel(writer, index=False, sheet_name=sh_c100)
                c190_x.to_excel(writer, index=False, sheet_name=sh_c190)
                _xlsx_reforcar_detalhe_texto_e_numero(
                    writer, sh_c190, c190_x, _COL_DETALHE_C190
                )
                if len(c100_x) > 0:
                    _xlsx_tabela_sobre_df_escrito(writer.sheets[sh_c100], c100_x)
                if len(c190_x) > 0:
                    _xlsx_tabela_sobre_df_escrito(writer.sheets[sh_c190], c190_x)
                _escrever_somatorio_cfop_unica(
                    writer,
                    tab_c,
                    nome_comp_c,
                    "C190",
                    rotulo_lado,
                    sh_c190,
                    c190_x,
                )
            if tem_d:
                tab_d = tabela_somatorio_cfop_estilo_excel(bl["D190"])
                d100_x.to_excel(writer, index=False, sheet_name=sh_d100)
                d190_x.to_excel(writer, index=False, sheet_name=sh_d190)
                _xlsx_reforcar_detalhe_texto_e_numero(
                    writer, sh_d190, d190_x, _COL_DETALHE_D190
                )
                if len(d100_x) > 0:
                    _xlsx_tabela_sobre_df_escrito(writer.sheets[sh_d100], d100_x)
                if len(d190_x) > 0:
                    _xlsx_tabela_sobre_df_escrito(writer.sheets[sh_d190], d190_x)
                _escrever_somatorio_cfop_unica(
                    writer,
                    tab_d,
                    nome_comp_d,
                    "D190",
                    rotulo_lado,
                    sh_d190,
                    d190_x,
                )

            for name in writer.sheets:
                writer.sheets[name].set_column(0, 40, 14)
            if tem_c:
                _xlsx_largura_coluna_chave(writer, sh_c100, c100_x, _COL_C100_PLAN)
                _xlsx_largura_coluna_chave(writer, sh_c190, c190_x, _COL_DETALHE_C190)
                _xlsx_largura_coluna_cnpj_emitente_nfe(writer, sh_c190, c190_x)
            if tem_d:
                _xlsx_largura_coluna_chave(writer, sh_d100, d100_x, _COL_C100_PLAN)
                _xlsx_largura_coluna_chave(writer, sh_d190, d190_x, _COL_DETALHE_D190)

            _xlsx_escrever_blocos_apur_difal_st(writer, bl, None)

            _xlsx_aplicar_abas_rosa(writer)
        output.seek(0)
        output = _xlsx_workbook_posprocesso_liberar(output)
        return output, "Sucesso"
    except Exception as e:
        return None, f"Erro: {str(e)}"


def _reg_id_de_nome_aba(nome: str) -> str:
    nome = (nome or "").strip()
    if not nome:
        return nome
    parte = re.split(r"\s*-\s*", nome, maxsplit=1)[0].strip()
    return parte.upper()


def _decodificar_txt(file_obj) -> str:
    file_obj.seek(0)
    raw = file_obj.read()
    if isinstance(raw, str):
        return raw
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace")


def _partes_linha_sped(line: str) -> list[str] | None:
    line = line.strip()
    if not line or line.startswith("#"):
        return None
    parts = line.split("|")
    if parts and parts[0] == "":
        parts = parts[1:]
    if parts and parts[-1] == "":
        parts = parts[:-1]
    if not parts or not parts[0].strip():
        return None
    return parts


def _reforcar_valores_sped_txt_por_posicao(texto: str, blocos: dict[str, pd.DataFrame]) -> None:
    """
    Sobrescreve VL_DOC / totais C190 (e D100/D190) com o split | posicional do Guia.
    O EFDReader fixa N colunas e trunca linhas; arquivos com campos a mais/menos deslocam
    VL_* — o parser por posição replica o comportamento típico da ferramenta de referência.
    """
    c100_vl: list[str] = []
    c190_src: list[dict[str, str]] = []
    d100_vl: list[str] = []
    d190_src: list[dict[str, str]] = []

    for line in texto.splitlines():
        parts = _partes_linha_sped(line)
        if not parts:
            continue
        reg = parts[0].strip().upper()
        if reg == "C100":
            c100_vl.append(parts[11] if len(parts) > 11 else "")
        elif reg == "C190":
            c190_src.append(
                {
                    "VL_OPR": parts[4] if len(parts) > 4 else "",
                    "VL_BC_ICMS": parts[5] if len(parts) > 5 else "",
                    "VL_ICMS": parts[6] if len(parts) > 6 else "",
                    "VL_BC_ICMS_ST": parts[7] if len(parts) > 7 else "",
                    "VL_ICMS_ST": parts[8] if len(parts) > 8 else "",
                    "VL_IPI": parts[10] if len(parts) > 10 else "",
                }
            )
        elif reg == "D100":
            d100_vl.append(parts[14] if len(parts) > 14 else "")
        elif reg == "D190":
            d190_src.append(
                {
                    "VL_OPR": parts[4] if len(parts) > 4 else "",
                    "VL_BC_ICMS": parts[5] if len(parts) > 5 else "",
                    "VL_ICMS": parts[6] if len(parts) > 6 else "",
                    "VL_RED_BC": parts[7] if len(parts) > 7 else "",
                }
            )

    def _aplica_col(df: pd.DataFrame, valores: list[str], col: str) -> pd.DataFrame:
        if df.empty or col not in df.columns or not valores:
            return df
        n = min(len(df), len(valores))
        out = df.copy()
        j = out.columns.get_loc(col)
        if isinstance(j, slice):
            return df
        for i in range(n):
            out.iat[i, j] = valores[i]
        return out

    def _aplica_map(df: pd.DataFrame, linhas: list[dict[str, str]]) -> pd.DataFrame:
        if df.empty or not linhas:
            return df
        n = min(len(df), len(linhas))
        out = df.copy()
        for i in range(n):
            for col, v in linhas[i].items():
                if col not in out.columns:
                    continue
                j = out.columns.get_loc(col)
                if isinstance(j, slice):
                    continue
                out.iat[i, j] = v
        return out

    if "C100" in blocos:
        blocos["C100"] = _aplica_col(blocos["C100"], c100_vl, "VL_DOC")
    if "C190" in blocos:
        blocos["C190"] = _aplica_map(blocos["C190"], c190_src)
    if "D100" in blocos:
        blocos["D100"] = _aplica_col(blocos["D100"], d100_vl, "VL_DOC")
    if "D190" in blocos:
        blocos["D190"] = _aplica_map(blocos["D190"], d190_src)


def _listar_regs_txt(file_obj) -> list[str]:
    text = _decodificar_txt(file_obj)
    regs = set()
    for line in text.splitlines():
        parts = _partes_linha_sped(line)
        if parts:
            regs.add(parts[0].strip().upper())
    file_obj.seek(0)
    return sorted(regs)


def _parse_sped_txt(file_obj) -> dict[str, pd.DataFrame]:
    text = _decodificar_txt(file_obj)
    buckets: dict[str, list[dict]] = {}
    for line in text.splitlines():
        parts = _partes_linha_sped(line)
        if not parts:
            continue
        reg = parts[0].strip().upper()
        row = {"REG": reg}
        for i, val in enumerate(parts[1:], start=2):
            row[f"C{i:02d}"] = val
        buckets.setdefault(reg, []).append(row)
    out: dict[str, pd.DataFrame] = {}
    for reg, rows in buckets.items():
        df = pd.DataFrame(rows)
        df = df.fillna("").astype(str)
        out[reg] = _aplicar_aliases_reg(reg, df)
    file_obj.seek(0)
    return out


def _aplicar_aliases_reg(reg: str, df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for src, dst in ALIASES_POR_REG.get(reg, []):
        if src in df.columns and dst not in df.columns:
            df[dst] = df[src]
    return df


def _listar_abas_excel(file_obj):
    file_obj.seek(0)
    nomes = pd.ExcelFile(file_obj).sheet_names
    file_obj.seek(0)
    return nomes


def listar_registros_arquivo(file_obj, filename: str) -> list[str]:
    fn = (filename or "").lower()
    if fn.endswith(".txt"):
        return _listar_regs_txt(file_obj)
    return sorted({_reg_id_de_nome_aba(s) for s in _listar_abas_excel(file_obj)})


def _carregar_mapa_abas(file_obj, filename: str) -> dict[str, pd.DataFrame]:
    fn = (filename or "").lower()
    if fn.endswith(".txt"):
        return _parse_sped_txt(file_obj)
    file_obj.seek(0)
    xl = pd.ExcelFile(file_obj)
    out: dict[str, pd.DataFrame] = {}
    for sh in xl.sheet_names:
        key = _reg_id_de_nome_aba(sh)
        file_obj.seek(0)
        df = pd.read_excel(file_obj, sheet_name=sh, dtype=str)
        df = _aplicar_aliases_reg(key, df)
        if key in out:
            out[key] = pd.concat([out[key], df], ignore_index=True)
        else:
            out[key] = df
    file_obj.seek(0)
    return out


def _normalizar_coluna_chave(df: pd.DataFrame, chave: str) -> pd.DataFrame:
    df = df.copy()
    s = df[chave].astype(str).str.strip()
    if chave == "CNPJ":
        s = s.str.replace(r"\D", "", regex=True)
    df[chave] = s
    return df


def _detectar_chave(df_cliente: pd.DataFrame, df_nascel: pd.DataFrame):
    c1, c2 = set(df_cliente.columns), set(df_nascel.columns)
    for k in CHAVES_CANDIDATAS:
        if k in c1 and k in c2:
            return k
    return None


def _fingerprint_linha(row, common_cols):
    return tuple(_norm_val(row[c]) for c in common_cols)


def _comparar_por_chave(df_cliente, df_nascel, chave: str):
    df_cliente = _normalizar_coluna_chave(df_cliente, chave)
    df_nascel = _normalizar_coluna_chave(df_nascel, chave)

    k_cliente = set(df_cliente[chave].dropna())
    k_cliente.discard("")
    k_nascel = set(df_nascel[chave].dropna())
    k_nascel.discard("")

    so_cliente = df_cliente[~df_cliente[chave].isin(k_nascel)].copy()
    so_cliente["TIPO_DIFERENCA"] = f"Só no {ROTULO_CLIENTE}"

    so_nascel = df_nascel[~df_nascel[chave].isin(k_cliente)].copy()
    so_nascel["TIPO_DIFERENCA"] = f"Só no {ROTULO_NASCEL}"

    dup_cliente = df_cliente[df_cliente.duplicated(subset=[chave], keep=False)]
    dup_nascel = df_nascel[df_nascel.duplicated(subset=[chave], keep=False)]

    d_cliente = df_cliente.drop_duplicates(subset=[chave], keep="first")
    d_nascel = df_nascel.drop_duplicates(subset=[chave], keep="first")

    common_cols = sorted(set(d_cliente.columns) & set(d_nascel.columns) - {chave})
    merged = d_cliente.merge(
        d_nascel, on=chave, how="inner", suffixes=(SUF_CLIENTE, SUF_NASCEL)
    )

    def divergencias_por_linha(row):
        difs = []
        for col in common_cols:
            c_cli, c_nas = f"{col}{SUF_CLIENTE}", f"{col}{SUF_NASCEL}"
            if c_cli not in row.index and c_nas not in row.index:
                continue
            if c_cli not in row.index:
                difs.append(f"{col} (só no {ROTULO_CLIENTE})")
                continue
            if c_nas not in row.index:
                difs.append(f"{col} (só no {ROTULO_NASCEL})")
                continue
            if _norm_val(row[c_cli]) != _norm_val(row[c_nas]):
                difs.append(col)
        return ", ".join(difs)

    merged["COLUNAS_DIVERGENTES"] = merged.apply(divergencias_por_linha, axis=1)
    com_divergencia = merged[merged["COLUNAS_DIVERGENTES"] != ""].copy()
    identicos = len(merged) - len(com_divergencia)

    return {
        "modo": f"chave:{chave}",
        "so_cliente": so_cliente,
        "so_nascel": so_nascel,
        "com_divergencia": com_divergencia,
        "dup_cliente": dup_cliente,
        "dup_nascel": dup_nascel,
        "n_so_cliente": len(so_cliente),
        "n_so_nascel": len(so_nascel),
        "n_diverg": len(com_divergencia),
        "n_identicos": identicos,
        "n_dup_cliente": len(dup_cliente),
        "n_dup_nascel": len(dup_nascel),
    }


def _comparar_por_linha(df_cliente, df_nascel):
    common_cols = sorted(set(df_cliente.columns) & set(df_nascel.columns))
    if not common_cols:
        return {
            "modo": "linha (sem colunas comuns)",
            "so_cliente": pd.DataFrame(),
            "so_nascel": pd.DataFrame(),
            "com_divergencia": pd.DataFrame(),
            "dup_cliente": pd.DataFrame(),
            "dup_nascel": pd.DataFrame(),
            "n_so_cliente": 0,
            "n_so_nascel": 0,
            "n_diverg": 0,
            "n_identicos": 0,
            "n_dup_cliente": 0,
            "n_dup_nascel": 0,
            "obs": "Nenhuma coluna comum nesta aba.",
        }

    def fp_series(df):
        return df.apply(lambda r: _fingerprint_linha(r, common_cols), axis=1)

    fp_c = fp_series(df_cliente)
    fp_n = fp_series(df_nascel)

    dup_cliente = df_cliente[fp_c.duplicated(keep=False)].copy()
    dup_nascel = df_nascel[fp_n.duplicated(keep=False)].copy()

    ca = Counter(fp_c.tolist())
    cb = Counter(fp_n.tolist())
    diff_c = ca - cb
    diff_n = cb - ca
    n_pareados = sum((ca & cb).values())

    rem_c = diff_c.copy()
    so_cliente_rows = []
    for idx, row in df_cliente.iterrows():
        fp = fp_c.loc[idx]
        if rem_c[fp] > 0:
            rem_c[fp] -= 1
            so_cliente_rows.append(row)
    so_cliente = (
        pd.DataFrame(so_cliente_rows)
        if so_cliente_rows
        else pd.DataFrame(columns=df_cliente.columns)
    )
    if not so_cliente.empty:
        so_cliente = so_cliente.copy()
        so_cliente["TIPO_DIFERENCA"] = f"Só no {ROTULO_CLIENTE} (linha inteira)"

    rem_n = diff_n.copy()
    so_nascel_rows = []
    for idx, row in df_nascel.iterrows():
        fp = fp_n.loc[idx]
        if rem_n[fp] > 0:
            rem_n[fp] -= 1
            so_nascel_rows.append(row)
    so_nascel = (
        pd.DataFrame(so_nascel_rows)
        if so_nascel_rows
        else pd.DataFrame(columns=df_nascel.columns)
    )
    if not so_nascel.empty:
        so_nascel = so_nascel.copy()
        so_nascel["TIPO_DIFERENCA"] = f"Só no {ROTULO_NASCEL} (linha inteira)"

    return {
        "modo": "linha (todas as colunas comuns)",
        "so_cliente": so_cliente,
        "so_nascel": so_nascel,
        "com_divergencia": pd.DataFrame(),
        "dup_cliente": dup_cliente,
        "dup_nascel": dup_nascel,
        "n_so_cliente": len(so_cliente),
        "n_so_nascel": len(so_nascel),
        "n_diverg": 0,
        "n_identicos": n_pareados,
        "n_dup_cliente": len(dup_cliente),
        "n_dup_nascel": len(dup_nascel),
        "obs": "Sem chave NF-e/CT-e/CNPJ/COD_PART: compara o conteúdo das colunas comuns (linhas idênticas se cancelam); o excedente aparece como só Cliente ou só Nascel.",
    }


def _comparar_uma_aba(df_cliente, df_nascel, nome_aba: str):
    if df_cliente.empty and df_nascel.empty:
        return {
            "aba": nome_aba,
            "modo": "—",
            "so_cliente": pd.DataFrame(),
            "so_nascel": pd.DataFrame(),
            "com_divergencia": pd.DataFrame(),
            "dup_cliente": pd.DataFrame(),
            "dup_nascel": pd.DataFrame(),
            "n_so_cliente": 0,
            "n_so_nascel": 0,
            "n_diverg": 0,
            "n_identicos": 0,
            "n_dup_cliente": 0,
            "n_dup_nascel": 0,
            "obs": "Aba vazia nos dois arquivos.",
        }

    chave = _detectar_chave(df_cliente, df_nascel)
    if chave:
        r = _comparar_por_chave(df_cliente, df_nascel, chave)
        r["aba"] = nome_aba
        r["obs"] = ""
        return r

    r = _comparar_por_linha(df_cliente, df_nascel)
    r["aba"] = nome_aba
    return r


def comparar_dois_sped(
    file_sped_cliente,
    file_sped_nascel,
    abas_selecionadas: list,
    nome_arquivo_cliente: str = "",
    nome_arquivo_nascel: str = "",
):
    try:
        map_c = _carregar_mapa_abas(file_sped_cliente, nome_arquivo_cliente)
        map_n = _carregar_mapa_abas(file_sped_nascel, nome_arquivo_nascel)

        comuns = sorted(set(map_c.keys()) & set(map_n.keys()))
        comuns = [a for a in comuns if a in abas_selecionadas]
        if not comuns:
            return None, "Nenhum registro em comum (ou nenhum selecionado)."

        linhas_resumo = []
        all_so_c, all_so_n, all_div, all_dup_c, all_dup_n = [], [], [], [], []

        for aba in comuns:
            df_c = map_c[aba]
            df_n = map_n[aba]
            r = _comparar_uma_aba(df_c, df_n, aba)

            if not r["so_cliente"].empty:
                t = r["so_cliente"].copy()
                t.insert(0, "ABA_SPED", aba)
                all_so_c.append(t)
            if not r["so_nascel"].empty:
                t = r["so_nascel"].copy()
                t.insert(0, "ABA_SPED", aba)
                all_so_n.append(t)
            if not r["com_divergencia"].empty:
                t = r["com_divergencia"].copy()
                t.insert(0, "ABA_SPED", aba)
                all_div.append(t)
            if not r["dup_cliente"].empty:
                t = r["dup_cliente"].copy()
                t.insert(0, "ABA_SPED", aba)
                all_dup_c.append(t)
            if not r["dup_nascel"].empty:
                t = r["dup_nascel"].copy()
                t.insert(0, "ABA_SPED", aba)
                all_dup_n.append(t)

            linhas_resumo.append(
                {
                    "ABA_SPED": aba,
                    "MODO_COMPARACAO": r["modo"],
                    f"Só_{ROTULO_CLIENTE.replace(' ', '_')}": r["n_so_cliente"],
                    f"Só_{ROTULO_NASCEL.replace(' ', '_')}": r["n_so_nascel"],
                    "Com_divergencia_campos": r["n_diverg"],
                    "Pareados_ou_identicos": r["n_identicos"],
                    f"Dup_{ROTULO_CLIENTE.replace(' ', '_')}": r["n_dup_cliente"],
                    f"Dup_{ROTULO_NASCEL.replace(' ', '_')}": r["n_dup_nascel"],
                    "Observacao": r.get("obs", ""),
                }
            )

        resumo = pd.DataFrame(linhas_resumo)
        totais = pd.DataFrame(
            [
                {
                    "Indicador": "Total de registros (blocos) comparados",
                    "Quantidade": len(comuns),
                },
                {
                    "Indicador": f"Soma linhas só {ROTULO_CLIENTE}",
                    "Quantidade": int(resumo[f"Só_{ROTULO_CLIENTE.replace(' ', '_')}"].sum()),
                },
                {
                    "Indicador": f"Soma linhas só {ROTULO_NASCEL}",
                    "Quantidade": int(resumo[f"Só_{ROTULO_NASCEL.replace(' ', '_')}"].sum()),
                },
                {
                    "Indicador": "Soma linhas com divergência de campos (modo chave)",
                    "Quantidade": int(resumo["Com_divergencia_campos"].sum()),
                },
            ]
        )

        def _concat_or_empty(parts, cols_msg):
            if not parts:
                return pd.DataFrame(columns=["ABA_SPED"] if cols_msg else [])
            return pd.concat(parts, ignore_index=True)

        out_so_c = _concat_or_empty(all_so_c, True)
        out_so_n = _concat_or_empty(all_so_n, True)
        out_div = _concat_or_empty(all_div, True)
        out_dup_c = _concat_or_empty(all_dup_c, True)
        out_dup_n = _concat_or_empty(all_dup_n, True)

        onde_agir = montar_onde_agir_sped(out_div, out_so_c, out_so_n)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            onde_agir.to_excel(writer, index=False, sheet_name="ONDE_AGIR")
            totais.to_excel(writer, index=False, sheet_name="Resumo_Geral")
            resumo.to_excel(writer, index=False, sheet_name="Resumo_por_aba")
            out_so_c.to_excel(writer, index=False, sheet_name="So_Cliente")
            out_so_n.to_excel(writer, index=False, sheet_name="So_Nascel")
            out_div.to_excel(writer, index=False, sheet_name="Campos_divergentes")
            if not out_dup_c.empty:
                out_dup_c.to_excel(writer, index=False, sheet_name="Dup_Cliente")
            if not out_dup_n.empty:
                out_dup_n.to_excel(writer, index=False, sheet_name="Dup_Nascel")

            if "ONDE_AGIR" in writer.sheets:
                w_oa = writer.sheets["ONDE_AGIR"]
                w_oa.set_column(0, 0, 6)
                w_oa.set_column(1, 3, 16)
                w_oa.set_column(4, 4, 88)

            for name, sheet in writer.sheets.items():
                if name == "ONDE_AGIR":
                    continue
                sheet.set_column(0, 60, 16)

            _xlsx_aplicar_abas_rosa(writer)
        output.seek(0)
        output = _xlsx_workbook_posprocesso_liberar(output)
        return output, "Sucesso"
    except Exception as e:
        return None, f"Erro: {str(e)}"
